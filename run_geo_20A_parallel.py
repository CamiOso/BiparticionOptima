"""Geometric paralelo (4 núcleos) para las 6 filas pendientes de 20A-Elementos.
Usa fork de Linux: TPM cargada en proceso padre, workers la heredan sin pickle.
"""
import sys, time, gc, multiprocessing as mp
import numpy as np
import openpyxl

PROJECT = "/home/cami/Desktop/AnalisisDiseñoAlgoritmos/Proyecto/ProyectoAnalisis2026"
sys.path.insert(0, PROJECT)

EXCEL     = f"{PROJECT}/DatosPruebas2026_1.xlsx"
SHEET     = "20A-Elementos"
CSV       = f"{PROJECT}/src/.samples/N20A.npy"
SISTEMA   = "ABCDEFGHIJKLMNOPQRST"
ESTADO    = "10000000000000000000"
CONDICION = "1" * 20

GEO_COLS = {2: (7,8,9), 3: (13,14,15), 4: (19,20,21), 5: (25,26,27)}

FILAS = [
    (45, "ACEGIKMOQS",     "ABDEGHJKMNPQST"),
    (47, "ACEGIKMOQS",     "BDFHJLNPRT"),
    (52, "BDFHJLNPRT",     "ABDEGHJKMNPQST"),
    (53, "BDFHJLNPRT",     "ACEGIKMOQS"),
    (54, "BDFHJLNPRT",     "BDFHJLNPRT"),
    (55, "BCDEFGJKLMNO",   "BCDEFGHIJKLMNO"),
]

# TPM global — heredada por workers via fork (sin pickle)
_tpm = None

def to_mask(sub, sistema):
    return "".join("1" if c in sub else "0" for c in sistema)

def procesar_fila(args):
    from src.strategies.geometric import Geometric
    fila_idx, alc_letras, mec_letras = args
    alc_mask = to_mask(alc_letras, SISTEMA)
    mec_mask = to_mask(mec_letras, SISTEMA)
    n_max = max(len(alc_letras), len(mec_letras))
    geo = Geometric(_tpm)
    geo._usar_paralelizacion_costos = False  # sin hilos internos — ya somos un worker
    kdict = {}
    t_fila = time.perf_counter()
    for k in (2, 3, 4, 5):
        t0 = time.perf_counter()
        res = geo.aplicar_estrategia(
            estado_inicial=ESTADO,
            condicion=CONDICION,
            alcance=alc_mask,
            mecanismo=mec_mask,
            k=k,
        )
        elapsed = time.perf_counter() - t0
        kdict[k] = (str(res.particion) if hasattr(res, "particion") else "",
                    float(res.perdida), round(elapsed, 4))
        print(f"  [fila={fila_idx} k={k}] perdida={round(float(res.perdida),6)} t={round(elapsed,2)}s", flush=True)
        gc.collect()
    total = round(time.perf_counter() - t_fila, 1)
    print(f"  [fila={fila_idx}] LISTO {total}s  n_max={n_max}", flush=True)
    return fila_idx, kdict

def guardar(fila_idx, kdict):
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[SHEET]
    for k, (part, perd, t) in kdict.items():
        cp, cl, ct = GEO_COLS[k]
        ws.cell(row=fila_idx, column=cp, value=part)
        ws.cell(row=fila_idx, column=cl, value=perd)
        ws.cell(row=fila_idx, column=ct, value=t)
    wb.save(EXCEL)
    wb.close()
    print(f"[GUARDADO] fila {fila_idx}", flush=True)

if __name__ == "__main__":
    import os, multiprocessing as mp
    # 1 hilo BLAS por worker — evita contención de 16 hilos en 4 cores
    os.environ["OMP_NUM_THREADS"] = "1"
    os.environ["OPENBLAS_NUM_THREADS"] = "1"
    os.environ["MKL_NUM_THREADS"] = "1"
    mp.set_start_method("fork", force=True)

    print("Cargando TPM 20A (float32)...", flush=True)
    globals()["_tpm"] = np.load(CSV)
    print(f"TPM: {_tpm.shape}  | Lanzando Pool(4) — fork, sin pickle...", flush=True)
    t0 = time.perf_counter()
    with mp.Pool(4) as pool:  # sin initializer — workers heredan _tpm via fork
        for fila_idx, kdict in pool.imap_unordered(procesar_fila, FILAS):
            guardar(fila_idx, kdict)
    print(f"\nTodas las filas completadas en {round((time.perf_counter()-t0)/60,1)} min", flush=True)
