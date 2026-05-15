"""QNodos para UNA fila de 20A-Elementos. Lanzar varias instancias con &"""
import sys, time, gc, argparse
import numpy as np
import openpyxl

PROJECT = "/home/cami/Desktop/AnalisisDiseñoAlgoritmos/Proyecto/ProyectoAnalisis2026"
sys.path.insert(0, PROJECT)

EXCEL     = f"{PROJECT}/DatosPruebas2026_1.xlsx"
SHEET     = "20A-Elementos"
CSV       = f"{PROJECT}/src/.samples/N20A.npy"
SISTEMA   = "ABCDEFGHIJKLMNOPQRST"
ESTADO    = "10000000000000000000"
CONDICION = "1" * 20
# QNodos: particion, perdida, tiempo
Q_COLS = {2: (4,5,6), 3: (10,11,12), 4: (16,17,18), 5: (22,23,24)}

FILAS = {
    # alc=20 mec=20/19
    6:  ("ABCDEFGHIJKLMNOPQRST", "ABCDEFGHIJKLMNOPQRST"),
    7:  ("ABCDEFGHIJKLMNOPQRST", "ABCDEFGHIJKLMNOPQRS"),
    8:  ("ABCDEFGHIJKLMNOPQRST", "BCDEFGHIJKLMNOPQRST"),
    # alc=20 mec=18/14/10
    11: ("ABCDEFGHIJKLMNOPQRST", "ACEGIKMOQS"),
    12: ("ABCDEFGHIJKLMNOPQRST", "BDFHJLNPRT"),
    10: ("ABCDEFGHIJKLMNOPQRST", "ABDEGHJKMNPQST"),
    9:  ("ABCDEFGHIJKLMNOPQRST", "BCDEFGHIJKLMNOPQRS"),
    # alc=19 ABCDEFGHIJKLMNOPQRS mec=20/19
    13: ("ABCDEFGHIJKLMNOPQRS",  "ABCDEFGHIJKLMNOPQRST"),
    14: ("ABCDEFGHIJKLMNOPQRS",  "ABCDEFGHIJKLMNOPQRS"),
    15: ("ABCDEFGHIJKLMNOPQRS",  "BCDEFGHIJKLMNOPQRST"),
    # alc=19 ABCDEFGHIJKLMNOPQRS mec=18/14/10
    18: ("ABCDEFGHIJKLMNOPQRS",  "ACEGIKMOQS"),
    19: ("ABCDEFGHIJKLMNOPQRS",  "BDFHJLNPRT"),
    17: ("ABCDEFGHIJKLMNOPQRS",  "ABDEGHJKMNPQST"),
    16: ("ABCDEFGHIJKLMNOPQRS",  "BCDEFGHIJKLMNOPQRS"),
    # alc=19 BCDEFGHIJKLMNOPQRST mec=20/19
    20: ("BCDEFGHIJKLMNOPQRST",  "ABCDEFGHIJKLMNOPQRST"),
    21: ("BCDEFGHIJKLMNOPQRST",  "ABCDEFGHIJKLMNOPQRS"),
    22: ("BCDEFGHIJKLMNOPQRST",  "BCDEFGHIJKLMNOPQRST"),
    # alc=19 BCDEFGHIJKLMNOPQRST mec=18/14/10
    25: ("BCDEFGHIJKLMNOPQRST",  "ACEGIKMOQS"),
    26: ("BCDEFGHIJKLMNOPQRST",  "BDFHJLNPRT"),
    24: ("BCDEFGHIJKLMNOPQRST",  "ABDEGHJKMNPQST"),
    23: ("BCDEFGHIJKLMNOPQRST",  "BCDEFGHIJKLMNOPQRS"),
    # alc=18 BCDEFGHIJKLMNOPQRS mec=20/19
    27: ("BCDEFGHIJKLMNOPQRS", "ABCDEFGHIJKLMNOPQRST"),
    28: ("BCDEFGHIJKLMNOPQRS", "ABCDEFGHIJKLMNOPQRS"),
    29: ("BCDEFGHIJKLMNOPQRS", "BCDEFGHIJKLMNOPQRST"),
    # alc=18 ya hechos
    30: ("BCDEFGHIJKLMNOPQRS", "BCDEFGHIJKLMNOPQRS"),
    31: ("BCDEFGHIJKLMNOPQRS", "ABDEGHJKMNPQST"),
    32: ("BCDEFGHIJKLMNOPQRS", "ACEGIKMOQS"),
    33: ("BCDEFGHIJKLMNOPQRS", "BDFHJLNPRT"),
    # alc=14 mec=20/19
    34: ("ABDEGHJKMNPQST",     "ABCDEFGHIJKLMNOPQRST"),
    35: ("ABDEGHJKMNPQST",     "ABCDEFGHIJKLMNOPQRS"),
    36: ("ABDEGHJKMNPQST",     "BCDEFGHIJKLMNOPQRST"),
    37: ("ABDEGHJKMNPQST",     "BCDEFGHIJKLMNOPQRS"),
    44: ("ACEGIKMOQS",         "BCDEFGHIJKLMNOPQRS"),
    51: ("BDFHJLNPRT",         "BCDEFGHIJKLMNOPQRS"),
    # alc=10 mec=20/19
    41: ("ACEGIKMOQS",  "ABCDEFGHIJKLMNOPQRST"),
    42: ("ACEGIKMOQS",  "ABCDEFGHIJKLMNOPQRS"),
    43: ("ACEGIKMOQS",  "BCDEFGHIJKLMNOPQRST"),
    48: ("BDFHJLNPRT",  "ABCDEFGHIJKLMNOPQRST"),
    49: ("BDFHJLNPRT",  "ABCDEFGHIJKLMNOPQRS"),
    50: ("BDFHJLNPRT",  "BCDEFGHIJKLMNOPQRST"),
    # alc=10 pendientes
    45: ("ACEGIKMOQS",   "ABDEGHJKMNPQST"),
    47: ("ACEGIKMOQS",   "BDFHJLNPRT"),
    52: ("BDFHJLNPRT",   "ABDEGHJKMNPQST"),
    53: ("BDFHJLNPRT",   "ACEGIKMOQS"),
    54: ("BDFHJLNPRT",   "BDFHJLNPRT"),
    55: ("BCDEFGJKLMNO", "BCDEFGHIJKLMNO"),
}

LOCK = EXCEL + ".lock"

def to_mask(sub):
    return "".join("1" if c in sub else "0" for c in SISTEMA)

def guardar_k(fila_idx, k, part, perd, elapsed):
    import os, random, time as _t
    cp, cl, ct = Q_COLS[k]
    for _ in range(20):
        try:
            fd = os.open(LOCK, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.close(fd)
            break
        except FileExistsError:
            _t.sleep(random.uniform(0.5, 2.0))
    else:
        raise RuntimeError("No se pudo adquirir lock del Excel")
    try:
        wb = openpyxl.load_workbook(EXCEL)
        ws = wb[SHEET]
        ws.cell(row=fila_idx, column=cp, value=part)
        ws.cell(row=fila_idx, column=cl, value=perd)
        ws.cell(row=fila_idx, column=ct, value=round(elapsed, 4))
        wb.save(EXCEL)
        wb.close()
    finally:
        try:
            os.unlink(LOCK)
        except FileNotFoundError:
            pass

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("fila", type=int)
    parser.add_argument("--start-k", type=int, default=2, choices=[2, 3, 4, 5])
    args = parser.parse_args()
    fila_idx = args.fila
    alc_letras, mec_letras = FILAS[fila_idx]

    from src.estrategias.q_nodos import QNodos
    print(f"[fila={fila_idx}] Cargando TPM...", flush=True)
    tpm = np.load(CSV)
    qnodos = QNodos(tpm)

    alc_mask = to_mask(alc_letras)
    mec_mask = to_mask(mec_letras)
    n_max = max(len(alc_letras), len(mec_letras))
    print(f"[fila={fila_idx}] n_max={n_max} | {alc_letras} / {mec_letras} | start_k={args.start_k}", flush=True)

    t_fila = time.perf_counter()
    for k in (k for k in (2, 3, 4, 5) if k >= args.start_k):
        t0 = time.perf_counter()
        res = qnodos.aplicar_estrategia(
            estado_inicial=ESTADO, condicion=CONDICION,
            alcance=alc_mask, mecanismo=mec_mask, k=k,
        )
        elapsed = time.perf_counter() - t0
        part = str(res.particion) if hasattr(res, "particion") else ""
        perd = float(res.perdida)
        print(f"  [fila={fila_idx} k={k}] perdida={round(perd,6)} t={round(elapsed,2)}s", flush=True)
        guardar_k(fila_idx, k, part, perd, elapsed)
        print(f"  [fila={fila_idx} k={k}] GUARDADO", flush=True)
        gc.collect()

    total = round(time.perf_counter() - t_fila, 1)
    print(f"  [fila={fila_idx}] COMPLETO {total}s", flush=True)
