"""
Lee casos de la hoja '28A-Elementos' del Excel y genera casos_28a.json
(ordenados por |alc|×|mec| ascendente, igual que casos_27a.json).

Uso:
    python3 implementacion_julia/extraer_casos_28a.py
"""
import json, os
import openpyxl

PROJ  = os.path.dirname(os.path.abspath(__file__))
EXCEL = os.path.join(PROJ, "..", "DatosIBQNodos2026.xlsx")
OUT   = os.path.join(PROJ, "casos_28a.json")

wb = openpyxl.load_workbook(EXCEL, read_only=True)

if "28A-Elementos" not in wb.sheetnames:
    raise SystemExit("ERROR: No existe hoja '28A-Elementos' en el Excel.\n"
                     "Ejecuta primero: python3 scripts/crear_hoja_28A.py")

ws = wb["28A-Elementos"]

casos = []
for row in ws.iter_rows(min_row=6, values_only=True):
    prueba, alc, mec = row[0], row[1], row[2]
    if prueba is None or alc is None or mec is None:
        continue
    casos.append({
        "prueba": int(prueba),
        "alc":    str(alc).strip(),
        "mec":    str(mec).strip(),
    })

if not casos:
    raise SystemExit("ERROR: No se encontraron casos en '28A-Elementos' (filas desde la 6).\n"
                     "Llena las columnas A=#Prueba, B=Alcance, C=Mecanismo en el Excel.")

casos.sort(key=lambda c: len(c["alc"]) * len(c["mec"]))

with open(OUT, "w") as f:
    json.dump(casos, f, indent=2)

print(f"Extraídos {len(casos)} casos → {OUT}")
print(f"Rango de costo: {len(casos[0]['alc'])*len(casos[0]['mec'])} "
      f"→ {len(casos[-1]['alc'])*len(casos[-1]['mec'])}")
