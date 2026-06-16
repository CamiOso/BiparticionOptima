"""
Actualiza DatosIBQNodos2026.xlsx al terminar n=26:
  1. Escribe resultados Julia en hoja '26A-Elementos' (cols D=partición, E=φ, F=tiempo)
  2. Crea/actualiza hoja '26A-IBQNodos' con resumen de resultados
  3. Crea hoja '27A-Elementos' con plantilla vacía para n=27
  4. Crea hoja '27A-IBQNodos' vacía (lista para resultados futuros)

Uso:
    python3 actualizar_excel_final.py
"""
import json, os
import openpyxl

PROJ  = os.path.dirname(os.path.abspath(__file__))
EXCEL = os.path.join(PROJ, "..", "DatosIBQNodos2026.xlsx")
RES   = os.path.join(PROJ, "resultados_26a.json")

SISTEMA_26 = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
SISTEMA_27 = "ABCDEFGHIJKLMNOPQRSTUVWXYZ0"
ESTADO_26  = "1" + "0" * 25
ESTADO_27  = "1" + "0" * 26

# ── Cargar resultados n=26 ────────────────────────────────────────────────────
with open(RES) as f:
    data = json.load(f)
casos = data["casos"]
print(f"Resultados n=26: {len(casos)} casos")

# ── Abrir Excel ───────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL)

# ── 1. Actualizar 26A-Elementos ───────────────────────────────────────────────
ws26 = wb["26A-Elementos"]
for c in casos:
    fila = c.get("fila")
    phi  = c.get("phi")
    t    = c.get("t")
    part = c.get("particion", "")
    if fila is None:
        continue
    ws26.cell(row=fila, column=4, value=part)
    ws26.cell(row=fila, column=5, value=round(phi, 8) if phi is not None else "TIMEOUT")
    ws26.cell(row=fila, column=6, value=round(t, 3) if t is not None else "")
print("  ✓ 26A-Elementos actualizado")

# ── 2. Crear/actualizar hoja 26A-IBQNodos-Julia ───────────────────────────────
sheet_name = "26A-IBQNodos-Julia"
if sheet_name in wb.sheetnames:
    del wb[sheet_name]
ws_res = wb.create_sheet(sheet_name)
ws_res.append(["#Prueba", "Alcance", "Mecanismo", "φ (IBQNodos)", "Tiempo(s)", "Partición"])
for c in sorted(casos, key=lambda x: x.get("fila", 999)):
    prueba = c.get("fila", 5) - 5
    phi    = c.get("phi")
    t      = c.get("t")
    ws_res.append([
        prueba,
        c.get("alc_str", ""),
        c.get("mec_str", ""),
        round(phi, 8) if phi is not None else "TIMEOUT",
        round(t, 3) if t is not None else "",
        c.get("particion", ""),
    ])
print(f"  ✓ {sheet_name} creado ({len(casos)} filas)")

# ── 3. Crear 27A-Elementos (plantilla vacía) ──────────────────────────────────
if "27A-Elementos" not in wb.sheetnames:
    ws27 = wb.create_sheet("27A-Elementos")
    ws27.cell(row=1, column=1, value="Estado inicial")
    ws27.cell(row=1, column=2, value=ESTADO_27)
    ws27.cell(row=2, column=1, value="Sistema:")
    ws27.cell(row=2, column=2, value=SISTEMA_27)
    ws27.cell(row=3, column=1, value="Sistema Candidato:")
    ws27.cell(row=3, column=2, value=SISTEMA_27)
    ws27.cell(row=3, column=4, value="PRUEBAS  BIPARTICIONES")
    ws27.cell(row=4, column=4, value="IBQNodos")
    ws27.cell(row=5, column=1, value="#Prueba")
    ws27.cell(row=5, column=2, value="Alcance o Purview (t+1)")
    ws27.cell(row=5, column=3, value="Mecanismo(t)")
    ws27.cell(row=5, column=4, value="Partición")
    ws27.cell(row=5, column=5, value="Pérdida")
    ws27.cell(row=5, column=6, value="Tiempo")
    print("  ✓ 27A-Elementos creado (plantilla vacía — llena los casos)")
else:
    print("  · 27A-Elementos ya existe, no se modifica")

# ── 4. Crear 27A-IBQNodos (placeholder para resultados futuros) ───────────────
if "27A-IBQNodos" not in wb.sheetnames:
    ws27r = wb.create_sheet("27A-IBQNodos")
    ws27r.append(["#Prueba", "Alcance", "Mecanismo", "φ (IBQNodos)", "Tiempo(s)", "Partición"])
    ws27r.cell(row=2, column=1, value="(pendiente — correr main_ibqnodos27.jl)")
    print("  ✓ 27A-IBQNodos creado (placeholder)")
else:
    print("  · 27A-IBQNodos ya existe")

# ── Guardar ───────────────────────────────────────────────────────────────────
wb.save(EXCEL)
print(f"\nExcel guardado: {EXCEL}")
completados = sum(1 for c in casos if c.get("phi") is not None)
print(f"n=26 completados: {completados}/{len(casos)}")
print("\nPróximos pasos para n=27:")
print("  1. Llena los casos en la hoja '27A-Elementos' del Excel")
print("  2. python scripts/generar_N27A.py          # genera TPM 14.5GB (~8 min)")
print("  3. python implementacion_julia/extraer_casos_27a.py  # lee Excel → JSON")
print("  4. julia --project=implementacion_julia implementacion_julia/main_ibqnodos27.jl")
