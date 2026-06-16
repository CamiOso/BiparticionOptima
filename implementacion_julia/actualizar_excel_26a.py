"""
Actualiza DatosIBQNodos2026.xlsx con los resultados de Julia.
Ejecutar después de que main_ibqnodos.jl termine:
    python3 actualizar_excel_26a.py
"""
import json, os
import openpyxl

PROJ = os.path.dirname(os.path.abspath(__file__))
EXCEL = os.path.join(PROJ, "..", "DatosIBQNodos2026.xlsx")
RES   = os.path.join(PROJ, "resultados_26a.json")

with open(RES) as f:
    data = json.load(f)

casos = data["casos"]
print(f"Cargando {len(casos)} resultados...")

wb = openpyxl.load_workbook(EXCEL)

# Hoja 26A-Elementos: cols D=particion, E=phi, F=tiempo
ws1 = wb["26A-Elementos"]
for c in casos:
    fila = c.get("fila")
    phi  = c.get("phi")
    t    = c.get("t")
    part = c.get("particion", "")
    if fila is None:
        continue
    ws1.cell(row=fila, column=4, value=part)
    ws1.cell(row=fila, column=5, value=round(phi, 8) if phi is not None else "TIMEOUT")
    ws1.cell(row=fila, column=6, value=round(t, 3) if t is not None else "")

# Hoja 26A-IBQNodos-Julia: crear si no existe
if "26A-IBQNodos-Julia" not in wb.sheetnames:
    ws2 = wb.create_sheet("26A-IBQNodos-Julia")
    ws2.append(["#Prueba", "Alcance", "Mecanismo", "IBQNodos_φ", "IBQNodos_t(s)", "Partición_IBQNodos"])
else:
    ws2 = wb["26A-IBQNodos-Julia"]

for c in casos:
    prueba = c.get("fila", 5) - 5
    phi    = c.get("phi")
    t      = c.get("t")
    row    = prueba + 1
    ws2.cell(row=row, column=1, value=prueba)
    ws2.cell(row=row, column=2, value=c.get("alc_str", ""))
    ws2.cell(row=row, column=3, value=c.get("mec_str", ""))
    ws2.cell(row=row, column=4, value=round(phi, 8) if phi is not None else "TIMEOUT")
    ws2.cell(row=row, column=5, value=round(t, 3) if t is not None else "")
    ws2.cell(row=row, column=6, value=c.get("particion", ""))

wb.save(EXCEL)
print(f"Excel actualizado: {EXCEL}")
completados = sum(1 for c in casos if c.get("phi") is not None)
print(f"Completados: {completados}/{len(casos)}")
