"""
Actualiza DatosIBQNodos2026.xlsx con resultados de n=28:
  1. Escribe resultados Julia en hoja '28A-Elementos' (cols D=partición, E=φ, F=tiempo)
  2. Crea/actualiza hoja '28A-IBQNodos-Julia' con resumen de resultados

Uso:
    python3 implementacion_julia/actualizar_excel_28a.py
"""
import json, os
import openpyxl

PROJ  = os.path.dirname(os.path.abspath(__file__))
EXCEL = os.path.join(PROJ, "..", "DatosIBQNodos2026.xlsx")
RES   = os.path.join(PROJ, "resultados_28a.json")

with open(RES) as f:
    data = json.load(f)
casos = data["casos"]
print(f"Resultados n=28: {len(casos)} casos")

wb = openpyxl.load_workbook(EXCEL)

# ── 1. Actualizar 28A-Elementos ───────────────────────────────────────────────
ws28 = wb["28A-Elementos"]
for c in casos:
    fila = c.get("fila")
    phi  = c.get("phi")
    t    = c.get("t")
    part = c.get("particion", "")
    if fila is None:
        continue
    ws28.cell(row=fila, column=4, value=part)
    ws28.cell(row=fila, column=5, value=round(phi, 8) if phi is not None else "TIMEOUT")
    ws28.cell(row=fila, column=6, value=round(t, 3) if t is not None else "")
print("  ✓ 28A-Elementos actualizado")

# ── 2. Crear/actualizar hoja 28A-IBQNodos-Julia ───────────────────────────────
sheet_name = "28A-IBQNodos-Julia"
if sheet_name in wb.sheetnames:
    del wb[sheet_name]
ws_res = wb.create_sheet(sheet_name)
ws_res.append(["#Prueba", "Alcance", "Mecanismo", "φ (IBQNodos)", "Tiempo(s)", "Partición"])
for c in sorted(casos, key=lambda x: x.get("fila", 999)):
    prueba = c.get("fila", 6) - 5
    phi    = c.get("phi")
    t      = c.get("t")
    ws_res.append([
        prueba,
        c.get("alc_str", ""),
        c.get("mec_str", ""),
        round(phi, 8) if phi is not None else "TIMEOUT",
        round(t, 3) if t is not None else "",
        c.get("particion", ""),
    ])
print(f"  ✓ {sheet_name} creado ({len(casos)} filas)")

wb.save(EXCEL)
print(f"\nExcel guardado: {EXCEL}")
completados = sum(1 for c in casos if c.get("phi") is not None)
print(f"n=28 completados: {completados}/{len(casos)}")
