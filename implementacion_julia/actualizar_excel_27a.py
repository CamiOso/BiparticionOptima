"""
Actualiza DatosIBQNodos2026.xlsx con resultados de n=27:
  1. Escribe resultados Julia en hoja '27A-Elementos' (cols D=partición, E=φ, F=tiempo)
  2. Crea/actualiza hoja '27A-IBQNodos-Julia' con resumen de resultados

Uso:
    python3 implementacion_julia/actualizar_excel_27a.py
"""
import json, os
import openpyxl

PROJ  = os.path.dirname(os.path.abspath(__file__))
EXCEL = os.path.join(PROJ, "..", "DatosIBQNodos2026.xlsx")
RES   = os.path.join(PROJ, "resultados_27a.json")

# ── Cargar resultados n=27 ────────────────────────────────────────────────────
with open(RES) as f:
    data = json.load(f)
casos = data["casos"]
print(f"Resultados n=27: {len(casos)} casos")

# ── Abrir Excel ───────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL)

# ── 1. Actualizar 27A-Elementos ───────────────────────────────────────────────
ws27 = wb["27A-Elementos"]
for c in casos:
    fila = c.get("fila")
    phi  = c.get("phi")
    t    = c.get("t")
    part = c.get("particion", "")
    if fila is None:
        continue
    ws27.cell(row=fila, column=4, value=part)
    ws27.cell(row=fila, column=5, value=round(phi, 8) if phi is not None else "TIMEOUT")
    ws27.cell(row=fila, column=6, value=round(t, 3) if t is not None else "")
print("  ✓ 27A-Elementos actualizado")

# ── 2. Crear/actualizar hoja 27A-IBQNodos-Julia ───────────────────────────────
sheet_name = "27A-IBQNodos-Julia"
if sheet_name in wb.sheetnames:
    del wb[sheet_name]
ws_res = wb.create_sheet(sheet_name)
ws_res.append(["#Prueba", "Alcance", "Mecanismo", "φ (IBQNodos)", "Tiempo(s)", "Partición"])
for c in sorted(casos, key=lambda x: x.get("fila", 999)):
    prueba = c.get("fila", 6) - 5
    phi    = c.get("phi")
    t      = c.get("t")
    ws_res.append([
        prueba,
        c.get("alc_str", ""),
        c.get("mec_str", ""),
        round(phi, 8) if phi is not None else "TIMEOUT",
        round(t, 3) if t is not None else "",
        c.get("particion", ""),
    ])
print(f"  ✓ {sheet_name} creado ({len(casos)} filas)")

# ── Guardar ───────────────────────────────────────────────────────────────────
wb.save(EXCEL)
print(f"\nExcel guardado: {EXCEL}")
completados = sum(1 for c in casos if c.get("phi") is not None)
print(f"n=27 completados: {completados}/{len(casos)}")
