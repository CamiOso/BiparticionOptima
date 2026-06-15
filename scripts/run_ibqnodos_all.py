"""Benchmark completo: IBQNodos vs QNodos en todas las hojas de DatosPruebas2026_1.xlsx.

Corre IBQNodos en cada caso donde QNodos tiene phi de referencia y guarda
los resultados en DatosIBQNodos2026.xlsx con checkpoint JSON intermedio.

Uso:
    python run_ibqnodos_all.py                    # todas las hojas
    python run_ibqnodos_all.py --hojas 10A 15B   # solo esas hojas
    python run_ibqnodos_all.py --desde 10A --fila 5  # retomar desde fila 5 de 10A
"""

import argparse
import json
import re
import signal
import sys
import time
from pathlib import Path

import numpy as np
import openpyxl

TIMEOUT_CASO = 8000  # segundos máximos por caso (mec=25 necesita hasta ~6500s)


class _CasoTimeout(Exception):
    pass


def _sigalrm_handler(signum, frame):
    raise _CasoTimeout()


PROJECT = Path(__file__).parent.resolve()
sys.path.insert(0, str(PROJECT))

EXCEL_REF   = PROJECT / "DatosPruebas2026_1.xlsx"
EXCEL_OUT   = PROJECT / "DatosIBQNodos2026.xlsx"
CHECKPOINT  = PROJECT / "checkpoint_ibqnodos_all.json"

# Configuración por hoja: (n, sistema_letras, estado_binario, tpm_path, mmap)
HOJAS = {
    "10A": {
        "sheet":   "10A-Elementos",
        "n":       10,
        "sistema": "ABCDEFGHIJ",
        "estado":  "1000000000",
        "tpm":     PROJECT / "src/.samples/N10A.csv",
        "mmap":    False,
    },
    "15B": {
        "sheet":   "15B-Elementos",
        "n":       15,
        "sistema": "ABCDEFGHIJKLMNO",
        "estado":  "100000000000000",
        "tpm":     PROJECT / "src/.samples/N15B.csv",
        "mmap":    False,
    },
    "20A": {
        "sheet":   "20A-Elementos",
        "n":       20,
        "sistema": "ABCDEFGHIJKLMNOPQRST",
        "estado":  "10000000000000000000",
        "tpm":     PROJECT / "src/.samples/N20A.npy",
        "mmap":    False,
    },
    "22A": {
        "sheet":   "22A-Elementos",
        "n":       22,
        "sistema": "ABCDEFGHIJKLMNOPQRSTUV",
        "estado":  "1000000000000000000000",
        "tpm":     PROJECT / "src/.samples/N22A.npy",
        "mmap":    False,
    },
    "25A": {
        "sheet":   "25A-Elementos ",   # nombre tiene espacio en Excel
        "n":       25,
        "sistema": "ABCDEFGHIJKLMNOPQRSTUVWXY",
        "estado":  "1000000000000000000000000",
        "tpm":     PROJECT / "src/.samples/N25A.npy",
        "mmap":    True,
    },
}


def to_mask(letters: str, sistema: str) -> str:
    """Convierte letras (ej. 'ACEG') a máscara binaria (ej. '1010101...')."""
    s = set(letters.upper())
    return "".join("1" if c in s else "0" for c in sistema)


_RE_GRUPO_A = re.compile(r'\(M=\(([^)]*)\), A=\(([^)]*)\)\)')

def _parse_grupo_a(particion_str: str) -> "set | None":
    """Extrae grupo_a del string de partición: {(0,m)...} ∪ {(1,a)...}."""
    m = _RE_GRUPO_A.match(particion_str)
    if not m:
        return None

    def nums(s: str) -> list[int]:
        return [int(x.strip()) for x in s.split(",") if x.strip().lstrip("-").isdigit()]

    grupo_a: set = set()
    for mn in nums(m.group(1)):
        grupo_a.add((0, mn))
    for an in nums(m.group(2)):
        grupo_a.add((1, an))
    return grupo_a or None


def leer_casos(cfg: dict, solo_con_phi: bool = False) -> list[dict]:
    """Lee todos los casos de la hoja. Con solo_con_phi=True filtra los que tienen referencia QNodos."""
    wb = openpyxl.load_workbook(EXCEL_REF, data_only=True, read_only=True)
    ws = wb[cfg["sheet"]]
    sistema = cfg["sistema"]
    casos   = []
    for r in range(6, ws.max_row + 1):
        alc = ws.cell(r, 2).value
        mec = ws.cell(r, 3).value
        phi_ref = ws.cell(r, 5).value
        t_ref   = ws.cell(r, 6).value
        if alc is None:
            break
        if solo_con_phi and phi_ref is None:
            continue
        casos.append({
            "fila":    r,
            "alc_str": str(alc),
            "mec_str": str(mec),
            "alc_bin": to_mask(str(alc), sistema),
            "mec_bin": to_mask(str(mec), sistema),
            "phi_ref": float(phi_ref) if phi_ref is not None else None,
            "t_ref":   float(t_ref)   if t_ref   is not None else None,
        })
    wb.close()
    return casos


def cargar_tpm(cfg: dict) -> np.ndarray:
    tpm_path = cfg["tpm"]
    print(f"  Cargando TPM {tpm_path.name} ...", flush=True)
    t0 = time.perf_counter()
    if tpm_path.suffix == ".npy":
        try:
            tpm = np.load(tpm_path, mmap_mode="r" if cfg["mmap"] else None)
        except ValueError:
            # Raw float32 binary (no numpy header) — load via memmap
            n = cfg["n"]
            rows = 1 << n
            tpm = np.memmap(tpm_path, dtype=np.float32, mode="r", shape=(rows, n))
    else:
        tpm = np.genfromtxt(tpm_path, delimiter=",").astype(np.float32)
    print(f"  TPM cargada en {time.perf_counter()-t0:.1f}s | shape={tpm.shape}", flush=True)
    return tpm


def guardar_checkpoint(data: dict) -> None:
    CHECKPOINT.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def cargar_checkpoint() -> dict:
    if CHECKPOINT.exists():
        return json.loads(CHECKPOINT.read_text(encoding="utf-8"))
    return {}


def crear_excel_resultado(resultados: dict) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for hoja_key, casos in resultados.items():
        if not casos:
            continue
        cfg = HOJAS[hoja_key]
        ws  = wb.create_sheet(title=f"{hoja_key}-IBQNodos")

        # Encabezados
        ws.append([
            "#Fila", "Alcance", "Mecanismo",
            "QNodos_φ", "QNodos_t(s)",
            "IBQNodos_φ", "IBQNodos_t(s)",
            "Speedup", "Match_φ",
            "Partición_IBQNodos",
        ])

        for caso in casos:
            phi_ib  = caso.get("phi_ib")
            t_ib    = caso.get("t_ib")
            phi_ref = caso.get("phi_ref")
            t_ref   = caso.get("t_ref")
            if phi_ib is not None:
                speedup = round(t_ref / t_ib, 1) if (t_ref and t_ib) else None
                if phi_ref is None:
                    match = "—"
                elif abs(phi_ib - phi_ref) < 1e-5:
                    match = "✓"
                elif phi_ib < phi_ref:
                    match = "↓mejor"
                else:
                    match = "✗"
            else:
                speedup = None
                match   = "pendiente"
            ws.append([
                caso["fila"],
                caso["alc_str"],
                caso["mec_str"],
                round(phi_ref, 8) if phi_ref is not None else None,
                round(t_ref, 4)   if t_ref   is not None else None,
                round(phi_ib, 8)  if phi_ib  is not None else None,
                round(t_ib, 4)    if t_ib    is not None else None,
                speedup,
                match,
                caso.get("particion_ib", ""),
            ])

    wb.save(EXCEL_OUT)
    print(f"\nExcel guardado en {EXCEL_OUT}", flush=True)


def run_hoja(hoja_key: str, checkpoint: dict, desde_fila: int = 0,
             solo_con_phi: bool = False) -> list[dict]:
    cfg   = HOJAS[hoja_key]
    casos = leer_casos(cfg, solo_con_phi=solo_con_phi)
    print(f"\n{'='*60}", flush=True)
    print(f"  Hoja {hoja_key}: {len(casos)} casos totales", flush=True)

    if not casos:
        print("  Sin casos — saltando.", flush=True)
        return []

    tpm       = cargar_tpm(cfg)
    condicion = "1" * cfg["n"]
    estado    = cfg["estado"]

    from src.estrategias.ib_qnodos import IBQNodos
    estrategia = IBQNodos(tpm)

    resultados_hoja = list(checkpoint.get(hoja_key, []))
    done_filas      = {c["fila"] for c in resultados_hoja if "phi_ib" in c}

    # Seed cache: mec_bin -> grupo_a extraído del mejor resultado disponible
    seed_cache: dict[str, set] = {}
    for c in resultados_hoja:
        if c.get("phi_ib") is not None:
            ga = _parse_grupo_a(c.get("particion_ib", ""))
            if ga:
                seed_cache[c["mec_bin"]] = ga

    for i, caso in enumerate(casos):
        if caso["fila"] < desde_fila:
            continue
        if caso["fila"] in done_filas:
            print(f"  [{i+1}/{len(casos)}] fila={caso['fila']} — ya completado, saltando.", flush=True)
            continue

        n_alc = caso["alc_bin"].count("1")
        n_mec = caso["mec_bin"].count("1")
        phi_ref_str = f"phi_ref={caso['phi_ref']:.6f}" if caso["phi_ref"] is not None else "phi_ref=—"
        print(
            f"  [{i+1}/{len(casos)}] fila={caso['fila']} "
            f"alc={caso['alc_str']}({n_alc}) mec={caso['mec_str']}({n_mec}) "
            f"{phi_ref_str}",
            flush=True,
        )

        grupo_a_seed = seed_cache.get(caso["mec_bin"])
        if grupo_a_seed:
            print(f"    warm-start: mec_bin={caso['mec_bin'][:8]}... seed_size={len(grupo_a_seed)}", flush=True)

        try:
            signal.signal(signal.SIGALRM, _sigalrm_handler)
            signal.alarm(TIMEOUT_CASO)
            t0 = time.perf_counter()
            try:
                res = estrategia.aplicar_estrategia(
                    estado_inicial=estado,
                    condicion=condicion,
                    alcance=caso["alc_bin"],
                    mecanismo=caso["mec_bin"],
                    k=2,
                    grupo_a_seed=grupo_a_seed,
                )
            finally:
                signal.alarm(0)
            elapsed = time.perf_counter() - t0

            phi_ib  = float(res.perdida)
            phi_ref = caso["phi_ref"]
            speedup = round(caso["t_ref"] / elapsed, 1) if caso.get("t_ref") else None

            if phi_ref is None:
                match_str = "—"
            elif abs(phi_ib - phi_ref) < 1e-5:
                match_str = "✓"
            elif phi_ib < phi_ref:
                match_str = "↓mejor"
            else:
                match_str = "✗"

            ref_str = f"ref={phi_ref:.6f}" if phi_ref is not None else "ref=—"
            print(
                f"    φ={phi_ib:.6f} {ref_str} "
                f"{match_str}  t={elapsed:.1f}s"
                + (f"  speedup=×{speedup}" if speedup else ""),
                flush=True,
            )

            caso_resultado = {**caso, "phi_ib": phi_ib, "t_ib": round(elapsed, 3),
                              "particion_ib": str(res.particion)}
            # Actualizar seed_cache con la partición recién hallada
            ga_new = _parse_grupo_a(str(res.particion))
            if ga_new:
                seed_cache[caso["mec_bin"]] = ga_new
        except _CasoTimeout:
            elapsed = time.perf_counter() - t0
            print(f"    TIMEOUT (>{TIMEOUT_CASO}s)  t={elapsed:.0f}s", flush=True)
            caso_resultado = {**caso, "phi_ib": None, "t_ib": round(elapsed, 3),
                              "particion_ib": f"TIMEOUT >{TIMEOUT_CASO}s"}
        except Exception as exc:
            print(f"    ERROR: {exc}", flush=True)
            caso_resultado = {**caso, "phi_ib": None, "t_ib": None,
                              "particion_ib": f"ERROR: {exc}"}

        # Upsert en resultados_hoja
        idx = next((j for j, c in enumerate(resultados_hoja) if c["fila"] == caso["fila"]), None)
        if idx is not None:
            resultados_hoja[idx] = caso_resultado
        else:
            resultados_hoja.append(caso_resultado)

        checkpoint[hoja_key] = resultados_hoja
        guardar_checkpoint(checkpoint)

    return resultados_hoja


def main() -> None:
    parser = argparse.ArgumentParser(description="Benchmark IBQNodos vs QNodos en todas las hojas.")
    parser.add_argument("--hojas",  nargs="+", choices=list(HOJAS), default=list(HOJAS),
                        help="Hojas a procesar (default: todas).")
    parser.add_argument("--desde",  default=None,
                        help="Primera hoja desde la que empezar (útil para retomar).")
    parser.add_argument("--fila",   type=int, default=0,
                        help="Número de fila de referencia desde la que empezar en la primera hoja.")
    args = parser.parse_args()

    hojas_orden = list(HOJAS.keys())
    hojas_a_correr: list[str] = args.hojas

    if args.desde:
        if args.desde not in hojas_orden:
            print(f"Hoja desconocida: {args.desde}", file=sys.stderr)
            sys.exit(1)
        hojas_a_correr = hojas_orden[hojas_orden.index(args.desde):]

    checkpoint  = cargar_checkpoint()
    resultados  = {}

    # 10A y 15B solo tienen datos QNodos; 20A, 22A, 25A corren todos los casos
    SOLO_PHI = {"10A": True, "15B": True, "20A": True, "22A": False, "25A": False}

    t_total = time.perf_counter()
    for hoja_key in hojas_a_correr:
        desde_fila = args.fila if hoja_key == hojas_a_correr[0] else 0
        resultados[hoja_key] = run_hoja(
            hoja_key, checkpoint,
            desde_fila=desde_fila,
            solo_con_phi=SOLO_PHI.get(hoja_key, False),
        )

    crear_excel_resultado(resultados)

    elapsed_total = time.perf_counter() - t_total
    print(f"\nTotal: {elapsed_total/3600:.2f}h  ({elapsed_total:.0f}s)", flush=True)

    # Imprimir resumen
    print("\n" + "="*60)
    print("  RESUMEN POR HOJA")
    print("="*60)
    for hoja_key, casos in resultados.items():
        completados = [c for c in casos if c.get("phi_ib") is not None]
        con_ref     = [c for c in completados if c.get("phi_ref") is not None]
        matches     = [c for c in con_ref if abs(c["phi_ib"] - c["phi_ref"]) < 1e-5]
        better      = [c for c in con_ref if c["phi_ib"] - c["phi_ref"] < -1e-5]
        if not completados:
            print(f"  {hoja_key}: sin resultados")
            continue
        speedups = [c["t_ref"] / c["t_ib"] for c in completados if c.get("t_ref") and c.get("t_ib")]
        avg_sp   = sum(speedups) / len(speedups) if speedups else 0
        sin_ref  = len(completados) - len(con_ref)
        print(
            f"  {hoja_key}: {len(completados)}/{len(casos)} casos | "
            f"match={len(matches)}/{len(con_ref)} mejor={len(better)} sin_ref={sin_ref} | "
            f"speedup_promedio=×{avg_sp:.1f}"
        )


if __name__ == "__main__":
    main()
