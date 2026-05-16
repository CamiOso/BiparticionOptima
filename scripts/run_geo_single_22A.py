"""Geometric para UNA fila de 22A-Elementos. Lanzar varias instancias con &"""
import sys, time, gc, argparse
import numpy as np
import openpyxl

PROJECT = "/home/cami/Desktop/AnalisisDiseñoAlgoritmos/Proyecto/ProyectoAnalisis2026"
sys.path.insert(0, PROJECT)

EXCEL     = f"{PROJECT}/DatosPruebas2026_1.xlsx"
SHEET     = "22A-Elementos"
CSV       = f"{PROJECT}/src/.samples/N22A.npy"
SISTEMA   = "ABCDEFGHIJKLMNOPQRSTUV"
ESTADO    = "1000000000000000000000"
CONDICION = "1" * 22
GEO_COLS  = {2: (7,8,9), 3: (13,14,15), 4: (19,20,21), 5: (25,26,27)}

FILAS = {
    6:  ("ABCDEFGHIJKLMNOPQRSTUV", "ABCDEFGHIJKLMNOPQRSTUV"),
    7:  ("ABCDEFGHIJKLMNOPQRSTUV", "ABCDEFGHIJKLMNOPQRSTU"),
    8:  ("ABCDEFGHIJKLMNOPQRSTUV", "BCDEFGHIJKLMNOPQRSTUV"),
    9:  ("ABCDEFGHIJKLMNOPQRSTUV", "BCDEFGHIJKLMNOPQRSTU"),
    10: ("ABCDEFGHIJKLMNOPQRSTUV", "ABDEGHJKMNPQSTV"),
    11: ("ABCDEFGHIJKLMNOPQRSTUV", "ACEGIKMOQSU"),
    12: ("ABCDEFGHIJKLMNOPQRSTUV", "BDFHJLNPRTV"),
    13: ("ABCDEFGHIJKLMNOPQRSTU",  "ABCDEFGHIJKLMNOPQRSTUV"),
    14: ("ABCDEFGHIJKLMNOPQRSTU",  "ABCDEFGHIJKLMNOPQRSTU"),
    15: ("ABCDEFGHIJKLMNOPQRSTU",  "BCDEFGHIJKLMNOPQRSTUV"),
    16: ("ABCDEFGHIJKLMNOPQRSTU",  "BCDEFGHIJKLMNOPQRSTU"),
    17: ("ABCDEFGHIJKLMNOPQRSTU",  "ABDEGHJKMNPQSTV"),
    18: ("ABCDEFGHIJKLMNOPQRSTU",  "ACEGIKMOQSU"),
    19: ("ABCDEFGHIJKLMNOPQRSTU",  "BDFHJLNPRTV"),
    20: ("BCDEFGHIJKLMNOPQRSTUV",  "ABCDEFGHIJKLMNOPQRSTUV"),
    21: ("BCDEFGHIJKLMNOPQRSTUV",  "ABCDEFGHIJKLMNOPQRSTU"),
    22: ("BCDEFGHIJKLMNOPQRSTUV",  "BCDEFGHIJKLMNOPQRSTUV"),
    23: ("BCDEFGHIJKLMNOPQRSTUV",  "BCDEFGHIJKLMNOPQRSTU"),
    24: ("BCDEFGHIJKLMNOPQRSTUV",  "ABDEGHJKMNPQSTV"),
    25: ("BCDEFGHIJKLMNOPQRSTUV",  "ACEGIKMOQSU"),
    26: ("BCDEFGHIJKLMNOPQRSTUV",  "BDFHJLNPRTV"),
    27: ("BCDEFGHIJKLMNOPQRSTU",   "ABCDEFGHIJKLMNOPQRSTUV"),
    28: ("BCDEFGHIJKLMNOPQRSTU",   "ABCDEFGHIJKLMNOPQRSTU"),
    29: ("BCDEFGHIJKLMNOPQRSTU",   "BCDEFGHIJKLMNOPQRSTUV"),
    30: ("BCDEFGHIJKLMNOPQRSTU",   "BCDEFGHIJKLMNOPQRSTU"),
    31: ("BCDEFGHIJKLMNOPQRSTU",   "ABDEGHJKMNPQSTV"),
    32: ("BCDEFGHIJKLMNOPQRSTU",   "ACEGIKMOQSU"),
    33: ("BCDEFGHIJKLMNOPQRSTU",   "BDFHJLNPRTV"),
    34: ("ABDEGHJKMNPQSTV",        "ABCDEFGHIJKLMNOPQRSTUV"),
    35: ("ABDEGHJKMNPQSTV",        "ABCDEFGHIJKLMNOPQRSTU"),
    36: ("ABDEGHJKMNPQSTV",        "BCDEFGHIJKLMNOPQRSTUV"),
    37: ("ABDEGHJKMNPQSTV",        "BCDEFGHIJKLMNOPQRSTU"),
    38: ("ABDEGHJKMNPQSTV",        "ABDEGHJKMNPQSTV"),
    39: ("ABDEGHJKMNPQSTV",        "ACEGIKMOQSU"),
    40: ("ABDEGHJKMNPQSTV",        "BDFHJLNPRTV"),
    41: ("ACEGIKMOQSU",            "ABCDEFGHIJKLMNOPQRSTUV"),
    42: ("ACEGIKMOQSU",            "ABCDEFGHIJKLMNOPQRSTU"),
    43: ("ACEGIKMOQSU",            "BCDEFGHIJKLMNOPQRSTUV"),
    44: ("ACEGIKMOQSU",            "BCDEFGHIJKLMNOPQRSTU"),
    45: ("ACEGIKMOQSU",            "ABDEGHJKMNPQSTV"),
    46: ("ACEGIKMOQSU",            "ACEGIKMOQSU"),
    47: ("ACEGIKMOQSU",            "BDFHJLNPRTV"),
    48: ("BDFHJLNPRTV",            "ABCDEFGHIJKLMNOPQRSTUV"),
    49: ("BDFHJLNPRTV",            "ABCDEFGHIJKLMNOPQRSTU"),
    50: ("BDFHJLNPRTV",            "BCDEFGHIJKLMNOPQRSTUV"),
    51: ("BDFHJLNPRTV",            "BCDEFGHIJKLMNOPQRSTU"),
    52: ("BDFHJLNPRTV",            "ABDEGHJKMNPQSTV"),
    53: ("BDFHJLNPRTV",            "ACEGIKMOQSU"),
    54: ("BDFHJLNPRTV",            "BDFHJLNPRTV"),
    55: ("ACDEFGHIJKLMNOPQRST",    "ACDEFGHIJKLMNOPQRST"),
}

LOCK = EXCEL + ".lock"

def to_mask(sub):
    return "".join("1" if c in sub else "0" for c in SISTEMA)

def guardar_k(fila_idx, k, part, perd, elapsed):
    import os, random, time as _t
    cp, cl, ct = GEO_COLS[k]
    for _ in range(20):
        try:
            fd = os.open(LOCK, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.close(fd)
            break
        except FileExistsError:
            _t.sleep(random.uniform(0.5, 2.0))
    else:
        raise RuntimeError("No se pudo adquirir lock del Excel")
    try:
        wb = openpyxl.load_workbook(EXCEL)
        ws = wb[SHEET]
        ws.cell(row=fila_idx, column=cp, value=part)
        ws.cell(row=fila_idx, column=cl, value=perd)
        ws.cell(row=fila_idx, column=ct, value=round(elapsed, 4))
        wb.save(EXCEL)
        wb.close()
    finally:
        try:
            os.unlink(LOCK)
        except FileNotFoundError:
            pass

if __name__ == "__main__":
    import os
    os.environ.setdefault("OMP_NUM_THREADS", "1")
    os.environ.setdefault("OPENBLAS_NUM_THREADS", "1")

    parser = argparse.ArgumentParser()
    parser.add_argument("fila", type=int)
    parser.add_argument("--start-k", type=int, default=2, choices=[2, 3, 4, 5])
    args = parser.parse_args()
    fila_idx = args.fila
    alc_letras, mec_letras = FILAS[fila_idx]

    mec_n = len(mec_letras)
    if mec_n <= 13:
        _cn, _cs = 512, 1024
    elif mec_n <= 17:
        _cn, _cs = 128, 512
    elif mec_n <= 20:
        _cn, _cs = 64,  256
    else:
        _cn, _cs = 32,  128
    import src.modelos.nucleo.ncubo   as _ncubo_mod
    import src.modelos.nucleo.sistema as _sistema_mod
    _ncubo_mod._MAX_MEMO_NCUBE   = _cn
    _sistema_mod._MAX_MEMO_SISTEMA = _cs
    print(f"[fila={fila_idx}] cache ncube={_cn} sistema={_cs} (mec={mec_n})", flush=True)

    from src.strategies.geometric import Geometric
    print(f"[fila={fila_idx}] Cargando TPM...", flush=True)
    tpm = np.load(CSV, mmap_mode="r")
    geo = Geometric(tpm)
    geo._usar_paralelizacion_costos = False

    alc_mask = to_mask(alc_letras)
    mec_mask = to_mask(mec_letras)
    n_max = max(len(alc_letras), len(mec_letras))
    print(f"[fila={fila_idx}] n_max={n_max} | {alc_letras} / {mec_letras} | start_k={args.start_k}", flush=True)

    t_fila = time.perf_counter()
    for k in (k for k in (2, 3, 4, 5) if k >= args.start_k):
        t0 = time.perf_counter()
        res = geo.aplicar_estrategia(
            estado_inicial=ESTADO, condicion=CONDICION,
            alcance=alc_mask, mecanismo=mec_mask, k=k,
        )
        elapsed = time.perf_counter() - t0
        part = str(res.particion) if hasattr(res, "particion") else ""
        perd = float(res.perdida)
        print(f"  [fila={fila_idx} k={k}] perdida={round(perd,6)} t={round(elapsed,2)}s", flush=True)
        guardar_k(fila_idx, k, part, perd, elapsed)
        print(f"  [fila={fila_idx} k={k}] GUARDADO", flush=True)
        gc.collect()

    total = round(time.perf_counter() - t_fila, 1)
    print(f"  [fila={fila_idx}] COMPLETO {total}s", flush=True)
