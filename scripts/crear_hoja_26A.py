"""Crea la hoja 26A-Elementos en DatosPruebas2026_1.xlsx.

Replica exactamente el patrón de 25A-Elementos: 7 grupos de alcance × 7 mecanismos + 1 especial = 50 casos.

Sistema n=26: ABCDEFGHIJKLMNOPQRSTUVWXYZ
Estado inicial: 10000000000000000000000000

Patrones de subconjuntos (replicando 25A):
  mec/alc 26  = ABCDEFGHIJKLMNOPQRSTUVWXYZ  (completo)
  mec/alc 25a = ABCDEFGHIJKLMNOPQRSTUVWXY   (sin Z)
  mec/alc 25b = BCDEFGHIJKLMNOPQRSTUVWXYZ   (sin A)
  mec/alc 24  = BCDEFGHIJKLMNOPQRSTUVWXY    (sin A ni Z)
  mec/alc 18  = ABDEGHJKMNPQSTVWYZ          (tomar 2, saltar 1)
  mec/alc 13a = ACEGIKMOQSUWY               (posiciones impares)
  mec/alc 13b = BDFHJLNPRTVXZ              (posiciones pares, incluye Z)

Uso:
    python crear_hoja_26A.py
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

EXCEL = Path("DatosIBQNodos2026.xlsx")
NOMBRE_HOJA = "26A-Elementos"

SISTEMA  = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"   # 26 letras A-Z
ESTADO   = "1" + "0" * 25                  # 10000000000000000000000000

# ── Subconjuntos canónicos ──────────────────────────────────────────────
S26  = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"         # n=26 completo
S25a = "ABCDEFGHIJKLMNOPQRSTUVWXY"          # sin Z
S25b = "BCDEFGHIJKLMNOPQRSTUVWXYZ"          # sin A
S24  = "BCDEFGHIJKLMNOPQRSTUVWXY"           # sin A ni Z
# Patrón tomar-2-saltar-1: AB_DE_GH_JK_MN_PQ_ST_VW_YZ
S18  = "ABDEGHJKMNPQSTVWYZ"                 # 18 nodos
# Posiciones impares (A=1,C=3,...,Y=25): 13 nodos
S13a = "ACEGIKMOQSUWY"
# Posiciones pares  (B=2,D=4,...,Z=26): 13 nodos
S13b = "BDFHJLNPRTVXZ"
# Especial: 22 nodos (análogo al alc=21 de 25A)
S22  = "ACDEFGHIJKLMNOPQRSTVXZ"              # 22 nodos (sin B, U, W, Y)

# ── Grupos: (alcance, [mec1, mec2, ...]) ───────────────────────────────
ALCANCES = [
    (S26,  [S26, S25a, S25b, S24, S18, S13a, S13b]),
    (S25a, [S26, S25a, S25b, S24, S18, S13a, S13b]),
    (S25b, [S26, S25a, S25b, S24, S18, S13a, S13b]),
    (S24,  [S26, S25a, S25b, S24, S18, S13a, S13b]),
    (S18,  [S26, S25a, S25b, S24, S18, S13a, S13b]),
    (S13a, [S26, S25a, S25b, S24, S18, S13a, S13b]),
    (S13b, [S26, S25a, S25b, S24, S18, S13a, S13b]),
]
ESPECIAL = [(S22, S22)]   # caso especial: 1 caso

# ── Estilos ────────────────────────────────────────────────────────────
AZUL    = PatternFill("solid", fgColor="2E75B6")
CELESTE = PatternFill("solid", fgColor="BDD7EE")
GRIS    = PatternFill("solid", fgColor="F2F2F2")
BLANCO  = PatternFill("solid", fgColor="FFFFFF")

def bold_white(size=11):
    return Font(bold=True, color="FFFFFF", size=size)

def bold_blue(size=10):
    return Font(bold=True, color="1F3864", size=size)

center = Alignment(horizontal="center", vertical="center")
left   = Alignment(horizontal="left",   vertical="center")


def escribir_hoja(wb):
    if NOMBRE_HOJA in wb.sheetnames:
        del wb[NOMBRE_HOJA]
    ws = wb.create_sheet(NOMBRE_HOJA)

    # Anchos
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12

    # ── Metadatos ──────────────────────────────────────────────────────
    def meta(row, label, value):
        a = ws.cell(row, 1, label)
        a.font = bold_blue(); a.fill = CELESTE; a.alignment = left
        b = ws.cell(row, 2, value)
        b.font = Font(size=10); b.fill = CELESTE; b.alignment = left
        ws.merge_cells(f"B{row}:F{row}")

    meta(1, "Estado inicial", ESTADO)
    meta(2, "Sistema:", SISTEMA)
    # Fila 3: label + sistema (A-C) + pruebas header (D-F)
    ws.cell(3, 1, "Sistema Candidato:").font = bold_blue()
    ws.cell(3, 1).fill = CELESTE; ws.cell(3, 1).alignment = left
    ws.cell(3, 2, SISTEMA).font = Font(size=10)
    ws.cell(3, 2).fill = CELESTE; ws.cell(3, 2).alignment = left
    ws.merge_cells("B3:C3")

    c = ws.cell(3, 4, "PRUEBAS  BIPARTICIONES")
    c.font = bold_white(); c.fill = AZUL; c.alignment = center
    ws.merge_cells("D3:F3")

    # Fila 4: vacío + IBQNodos header
    c = ws.cell(4, 4, "IBQNodos")
    c.font = bold_white(); c.fill = AZUL; c.alignment = center
    ws.merge_cells("D4:F4")

    # ── Encabezados columnas ───────────────────────────────────────────
    headers = ["#Prueba", "Alcance o Purview (t+1)", "Mecanismo(t)",
               "Partición", "Pérdida", "Tiempo"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(5, col, h)
        c.font = bold_white(); c.fill = AZUL; c.alignment = center
    ws.row_dimensions[5].height = 20

    # ── Casos ──────────────────────────────────────────────────────────
    fila = 6
    prueba = 1

    for alc_str, mecanismos in ALCANCES:
        for i, mec_str in enumerate(mecanismos):
            bg = GRIS if i % 2 == 0 else BLANCO

            ws.cell(fila, 1, prueba).fill = bg
            ws.cell(fila, 2, alc_str).fill = bg
            ws.cell(fila, 3, mec_str).fill = bg
            # cols D, E, F vacías (se llenarán con IBQNodos)
            for col in [4, 5, 6]:
                ws.cell(fila, col).fill = bg

            for col in range(1, 7):
                ws.cell(fila, col).alignment = left
                ws.cell(fila, col).font = Font(size=10)

            fila += 1
            prueba += 1

    # Caso especial
    alc_str, mec_str = ESPECIAL[0]
    bg = CELESTE
    ws.cell(fila, 1, prueba).fill = bg; ws.cell(fila, 1).font = Font(size=10, bold=True)
    ws.cell(fila, 2, alc_str).fill = bg
    ws.cell(fila, 3, mec_str).fill = bg
    for col in [4, 5, 6]:
        ws.cell(fila, col).fill = bg

    print(f"Hoja '{NOMBRE_HOJA}' creada con {prueba} casos (filas 6–{fila}).")
    return ws


def main():
    wb = openpyxl.load_workbook(EXCEL)
    escribir_hoja(wb)
    wb.save(EXCEL)
    print(f"Guardado: {EXCEL.resolve()}")

    # Resumen de casos
    total = sum(len(m) for _, m in ALCANCES) + len(ESPECIAL)
    print(f"\nResumen 26A-Elementos:")
    print(f"  Total casos: {total}")
    for alc_str, mecanismos in ALCANCES:
        print(f"  alc={len(alc_str):2d} ({alc_str[:10]}...): {len(mecanismos)} mechs")
    print(f"  Especial alc={len(ESPECIAL[0][0])}: 1 caso")
    print(f"\nPatrones de subconjuntos usados:")
    for nombre, s in [("S26",S26),("S25a",S25a),("S25b",S25b),("S24",S24),
                       ("S18",S18),("S13a",S13a),("S13b",S13b),("S22",S22)]:
        print(f"  {nombre} ({len(s):2d} nodos): {s}")


if __name__ == "__main__":
    main()
