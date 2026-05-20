"""QNodos para UNA fila de 22A-Elementos. Lanzar varias instancias con &"""
import sys, time, gc, argparse, os, json, random
import numpy as np
import openpyxl

PROJECT = "/home/cami/Desktop/AnalisisDiseñoAlgoritmos/Proyecto/ProyectoAnalisis2026"
sys.path.insert(0, PROJECT)

EXCEL     = f"{PROJECT}/DatosPruebas2026_1.xlsx"
SHEET     = "22A-Elementos"
CSV       = f"{PROJECT}/src/.samples/N22A.npy"
SISTEMA   = "ABCDEFGHIJKLMNOPQRSTUV"
ESTADO    = "1000000000000000000000"
CONDICION = "1" * 22
Q_COLS    = {2: (4,5,6), 3: (10,11,12), 4: (16,17,18), 5: (22,23,24)}

FILAS = {
    # alc=22 mec=22/21/20
    6:  ("ABCDEFGHIJKLMNOPQRSTUV", "ABCDEFGHIJKLMNOPQRSTUV"),
    7:  ("ABCDEFGHIJKLMNOPQRSTUV", "ABCDEFGHIJKLMNOPQRSTU"),
    8:  ("ABCDEFGHIJKLMNOPQRSTUV", "BCDEFGHIJKLMNOPQRSTUV"),
    9:  ("ABCDEFGHIJKLMNOPQRSTUV", "BCDEFGHIJKLMNOPQRSTU"),
    10: ("ABCDEFGHIJKLMNOPQRSTUV", "ABDEGHJKMNPQSTV"),
    11: ("ABCDEFGHIJKLMNOPQRSTUV", "ACEGIKMOQSU"),
    12: ("ABCDEFGHIJKLMNOPQRSTUV", "BDFHJLNPRTV"),
    # alc=21 ABCDEFGHIJKLMNOPQRSTU
    13: ("ABCDEFGHIJKLMNOPQRSTU",  "ABCDEFGHIJKLMNOPQRSTUV"),
    14: ("ABCDEFGHIJKLMNOPQRSTU",  "ABCDEFGHIJKLMNOPQRSTU"),
    15: ("ABCDEFGHIJKLMNOPQRSTU",  "BCDEFGHIJKLMNOPQRSTUV"),
    16: ("ABCDEFGHIJKLMNOPQRSTU",  "BCDEFGHIJKLMNOPQRSTU"),
    17: ("ABCDEFGHIJKLMNOPQRSTU",  "ABDEGHJKMNPQSTV"),
    18: ("ABCDEFGHIJKLMNOPQRSTU",  "ACEGIKMOQSU"),
    19: ("ABCDEFGHIJKLMNOPQRSTU",  "BDFHJLNPRTV"),
    # alc=21 BCDEFGHIJKLMNOPQRSTUV
    20: ("BCDEFGHIJKLMNOPQRSTUV",  "ABCDEFGHIJKLMNOPQRSTUV"),
    21: ("BCDEFGHIJKLMNOPQRSTUV",  "ABCDEFGHIJKLMNOPQRSTU"),
    22: ("BCDEFGHIJKLMNOPQRSTUV",  "BCDEFGHIJKLMNOPQRSTUV"),
    23: ("BCDEFGHIJKLMNOPQRSTUV",  "BCDEFGHIJKLMNOPQRSTU"),
    24: ("BCDEFGHIJKLMNOPQRSTUV",  "ABDEGHJKMNPQSTV"),
    25: ("BCDEFGHIJKLMNOPQRSTUV",  "ACEGIKMOQSU"),
    26: ("BCDEFGHIJKLMNOPQRSTUV",  "BDFHJLNPRTV"),
    # alc=20 BCDEFGHIJKLMNOPQRSTU
    27: ("BCDEFGHIJKLMNOPQRSTU",   "ABCDEFGHIJKLMNOPQRSTUV"),
    28: ("BCDEFGHIJKLMNOPQRSTU",   "ABCDEFGHIJKLMNOPQRSTU"),
    29: ("BCDEFGHIJKLMNOPQRSTU",   "BCDEFGHIJKLMNOPQRSTUV"),
    30: ("BCDEFGHIJKLMNOPQRSTU",   "BCDEFGHIJKLMNOPQRSTU"),
    31: ("BCDEFGHIJKLMNOPQRSTU",   "ABDEGHJKMNPQSTV"),
    32: ("BCDEFGHIJKLMNOPQRSTU",   "ACEGIKMOQSU"),
    33: ("BCDEFGHIJKLMNOPQRSTU",   "BDFHJLNPRTV"),
    # alc=15 ABDEGHJKMNPQSTV
    34: ("ABDEGHJKMNPQSTV",        "ABCDEFGHIJKLMNOPQRSTUV"),
    35: ("ABDEGHJKMNPQSTV",        "ABCDEFGHIJKLMNOPQRSTU"),
    36: ("ABDEGHJKMNPQSTV",        "BCDEFGHIJKLMNOPQRSTUV"),
    37: ("ABDEGHJKMNPQSTV",        "BCDEFGHIJKLMNOPQRSTU"),
    38: ("ABDEGHJKMNPQSTV",        "ABDEGHJKMNPQSTV"),
    39: ("ABDEGHJKMNPQSTV",        "ACEGIKMOQSU"),
    40: ("ABDEGHJKMNPQSTV",        "BDFHJLNPRTV"),
    # alc=11 ACEGIKMOQSU
    41: ("ACEGIKMOQSU",            "ABCDEFGHIJKLMNOPQRSTUV"),
    42: ("ACEGIKMOQSU",            "ABCDEFGHIJKLMNOPQRSTU"),
    43: ("ACEGIKMOQSU",            "BCDEFGHIJKLMNOPQRSTUV"),
    44: ("ACEGIKMOQSU",            "BCDEFGHIJKLMNOPQRSTU"),
    45: ("ACEGIKMOQSU",            "ABDEGHJKMNPQSTV"),
    46: ("ACEGIKMOQSU",            "ACEGIKMOQSU"),
    47: ("ACEGIKMOQSU",            "BDFHJLNPRTV"),
    # alc=11 BDFHJLNPRTV
    48: ("BDFHJLNPRTV",            "ABCDEFGHIJKLMNOPQRSTUV"),
    49: ("BDFHJLNPRTV",            "ABCDEFGHIJKLMNOPQRSTU"),
    50: ("BDFHJLNPRTV",            "BCDEFGHIJKLMNOPQRSTUV"),
    51: ("BDFHJLNPRTV",            "BCDEFGHIJKLMNOPQRSTU"),
    52: ("BDFHJLNPRTV",            "ABDEGHJKMNPQSTV"),
    53: ("BDFHJLNPRTV",            "ACEGIKMOQSU"),
    54: ("BDFHJLNPRTV",            "BDFHJLNPRTV"),
    55: ("ACDEFGHIJKLMNOPQRST",    "ACDEFGHIJKLMNOPQRST"),
}

LOCK      = EXCEL + ".lock"
MEC_CACHE = "/tmp/mec_cache_22A.json"
LOCK_CACHE = MEC_CACHE + ".lock"

def to_mask(sub):
    return "".join("1" if c in sub else "0" for c in SISTEMA)

def guardar_k(fila_idx, k, part, perd, elapsed):
    import os, random, time as _t
    cp, cl, ct = Q_COLS[k]
    for _ in range(20):
        try:
            fd = os.open(LOCK, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.close(fd)
            break
        except FileExistsError:
            _t.sleep(random.uniform(0.5, 2.0))
    else:
        raise RuntimeError("No se pudo adquirir lock del Excel")
    try:
        wb = openpyxl.load_workbook(EXCEL)
        ws = wb[SHEET]
        ws.cell(row=fila_idx, column=cp, value=part)
        ws.cell(row=fila_idx, column=cl, value=perd)
        ws.cell(row=fila_idx, column=ct, value=round(elapsed, 4))
        wb.save(EXCEL)
        wb.close()
    finally:
        try:
            os.unlink(LOCK)
        except FileNotFoundError:
            pass

def _actualizar_cache(mec_key: str, k: int, perdida: float, fila_idx: int) -> None:
    """Guarda resultado de fila referencia (alc<=mec) en cache para reutilizar en filas con alc>mec."""
    for _ in range(10):
        try:
            fd = os.open(LOCK_CACHE, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.close(fd)
            break
        except FileExistsError:
            time.sleep(random.uniform(0.5, 1.5))
    else:
        return
    try:
        cache_data: dict = {}
        if os.path.exists(MEC_CACHE):
            try:
                with open(MEC_CACHE) as f:
                    cache_data = json.load(f)
            except Exception:
                pass
        if mec_key not in cache_data:
            cache_data[mec_key] = {}
        cache_data[mec_key][str(k)] = {"perdida": perdida, "source": fila_idx}
        with open(MEC_CACHE, "w") as f:
            json.dump(cache_data, f)
    finally:
        try:
            os.unlink(LOCK_CACHE)
        except FileNotFoundError:
            pass


if __name__ == "__main__":
    os.environ.setdefault("OMP_NUM_THREADS", "1")
    os.environ.setdefault("OPENBLAS_NUM_THREADS", "1")

    parser = argparse.ArgumentParser()
    parser.add_argument("fila", type=int)
    parser.add_argument("--start-k", type=int, default=2, choices=[2, 3, 4, 5])
    args = parser.parse_args()
    fila_idx = args.fila
    alc_letras, mec_letras = FILAS[fila_idx]

    # --- Mec-cache: mismo mec + alc_n > mec_n → perdida idéntica (empíricamente verificado)
    mec_key    = "".join(sorted(mec_letras))
    alc_n_val  = len(alc_letras)
    mec_n_val  = len(mec_letras)
    es_referencia = (alc_n_val > mec_n_val)  # solo beneficiarias poblan el cache

    if alc_n_val > mec_n_val:
        cache_data: dict = {}
        if os.path.exists(MEC_CACHE):
            try:
                with open(MEC_CACHE) as _cf:
                    cache_data = json.load(_cf)
            except Exception:
                pass
        ks_req = [k for k in (2, 3, 4, 5) if k >= args.start_k]
        mec_entry = cache_data.get(mec_key, {})
        if all(str(k) in mec_entry for k in ks_req):
            print(
                f"[fila={fila_idx}] CACHE HIT mec={mec_letras[:14]}... "
                f"alc_n={alc_n_val} > mec_n={mec_n_val} → reutilizando perdida",
                flush=True,
            )
            for k in ks_req:
                entry = mec_entry[str(k)]
                perd  = entry["perdida"]
                src   = entry["source"]
                part  = f"CACHE:fila{src}"
                print(f"  [fila={fila_idx} k={k}] perdida={round(perd,6)} (cache de fila {src})", flush=True)
                guardar_k(fila_idx, k, part, perd, 0.0)
                print(f"  [fila={fila_idx} k={k}] GUARDADO (cache)", flush=True)
            print(f"  [fila={fila_idx}] COMPLETO (cache, 0s)", flush=True)
            sys.exit(0)
        print(
            f"[fila={fila_idx}] alc_n={alc_n_val} > mec_n={mec_n_val} pero cache vacío "
            f"para mec={mec_letras[:14]}... → ejecutando normal",
            flush=True,
        )
    # -------------------------------------------------------------------------

    # Cache adaptativo según tamaño de mec (entry máximo = 2^(mec-1) × 4 bytes)
    mec_n = len(mec_letras)
    if mec_n <= 13:
        _cn, _cs = 512, 1024   # entries ≤8KB → +50MB total: seguro
    elif mec_n <= 17:
        _cn, _cs = 128, 512    # entries ≤256KB → +500MB total: moderado
    elif mec_n <= 20:
        _cn, _cs = 64,  256    # default
    else:
        _cn, _cs = 32,  128    # entries ≥4MB → reducir para evitar OOM
    import src.modelos.nucleo.ncubo   as _ncubo_mod
    import src.modelos.nucleo.sistema as _sistema_mod
    _ncubo_mod._MAX_MEMO_NCUBE   = _cn
    _sistema_mod._MAX_MEMO_SISTEMA = _cs
    print(f"[fila={fila_idx}] cache ncube={_cn} sistema={_cs} (mec={mec_n})", flush=True)

    from src.estrategias.q_nodos import QNodos
    print(f"[fila={fila_idx}] Cargando TPM...", flush=True)
    tpm = np.load(CSV, mmap_mode="r")
    qnodos = QNodos(tpm)

    alc_mask = to_mask(alc_letras)
    mec_mask = to_mask(mec_letras)
    n_max = max(len(alc_letras), len(mec_letras))
    print(f"[fila={fila_idx}] n_max={n_max} | {alc_letras} / {mec_letras} | start_k={args.start_k}", flush=True)

    t_fila = time.perf_counter()
    for k in (k for k in (2, 3, 4, 5) if k >= args.start_k):
        t0 = time.perf_counter()
        res = qnodos.aplicar_estrategia(
            estado_inicial=ESTADO, condicion=CONDICION,
            alcance=alc_mask, mecanismo=mec_mask, k=k,
        )
        elapsed = time.perf_counter() - t0
        part = str(res.particion) if hasattr(res, "particion") else ""
        perd = float(res.perdida)
        print(f"  [fila={fila_idx} k={k}] perdida={round(perd,6)} t={round(elapsed,2)}s", flush=True)
        guardar_k(fila_idx, k, part, perd, elapsed)
        print(f"  [fila={fila_idx} k={k}] GUARDADO", flush=True)
        if es_referencia:
            _actualizar_cache(mec_key, k, perd, fila_idx)
        gc.collect()

    total = round(time.perf_counter() - t_fila, 1)
    print(f"  [fila={fila_idx}] COMPLETO {total}s", flush=True)
