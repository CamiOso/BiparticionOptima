"""Genera DatosCircuito2026.xlsx a partir de DatosPruebas2026_1.xlsx.

Layout (39 columnas de datos):
  Cols 1-3  : #Prueba, Alcance, Mecanismo
  k=2 bloque (cols 4-12) : QNodos(4-6) | Geometric(7-9) | Circuit(10-12)
  k=3 bloque (cols 13-21): QNodos(13-15)| Geometric(16-18)| Circuit(19-21)
  k=4 bloque (cols 22-30): QNodos(22-24)| Geometric(25-27)| Circuit(28-30)
  k=5 bloque (cols 31-39): QNodos(31-33)| Geometric(34-36)| Circuit(37-39)

QNodos y Geometric se pre-rellenan desde DatosPruebas2026_1.xlsx.
Circuit permanece vacío (se llena con run_circuito_single.py).
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ORIGEN  = "DatosPruebas2026_1.xlsx"
DESTINO = "DatosCircuito2026.xlsx"

# (hoja_src, estado, sistema, tag_dst)
HOJAS = [
    ("10A-Elementos",  "1000000000",                     "ABCDEFGHIJ",                "10A"),
    ("15B-Elementos",  "100000000000000",                 "ABCDEFGHIJKLMNO",           "15B"),
    ("20A-Elementos",  "10000000000000000000",             "ABCDEFGHIJKLMNOPQRST",      "20A"),
    ("22A-Elementos",  "1000000000000000000000",           "ABCDEFGHIJKLMNOPQRSTUV",    "22A"),
    ("25A-Elementos ", "1000000000000000000000000",        "ABCDEFGHIJKLMNOPQRSTUVWXY", "25A"),
]

# Columnas en el Excel ORIGEN para cada k (1-indexed):
#   k=2: QNodos 4-6, Geometric 7-9
#   k=3: QNodos 10-12, Geometric 13-15
#   k=4: QNodos 16-18, Geometric 19-21
#   k=5: QNodos 22-24, Geometric 25-27
ORIG_COLS = {
    2: ((4, 5, 6),   (7, 8, 9)),
    3: ((10, 11, 12), (13, 14, 15)),
    4: ((16, 17, 18), (19, 20, 21)),
    5: ((22, 23, 24), (25, 26, 27)),
}

# Columna de inicio de cada bloque k en el DESTINO
DST_START = {2: 4, 3: 13, 4: 22, 5: 31}

# Colores
COLOR_QNODOS   = "D9EAD3"   # verde
COLOR_GEO      = "CFE2F3"   # azul
COLOR_CIRCUIT  = "E8D5F5"   # violeta
COLOR_HDR_K    = "434343"   # gris oscuro
COLOR_HDR_DARK = "666666"


def _fill(color):
    return PatternFill("solid", fgColor=color)


def _bold(size=11, color="000000"):
    return Font(bold=True, size=size, color=color)


def crear_hoja(wb_dst, ws_src, tag, estado, sistema):
    ws = wb_dst.create_sheet(title=tag)

    # Filas 1-2: metadatos
    ws.cell(1, 1, "Estado inicial").font = _bold()
    ws.cell(1, 2, estado)
    ws.cell(2, 1, "Sistema:").font = _bold()
    ws.cell(2, 2, sistema)
    ws.cell(3, 1, "Sistema Candidato:").font = _bold()
    ws.cell(3, 2, sistema)

    # Fila 3: headers de bloque k
    for k, cs in DST_START.items():
        lbl = f"PRUEBAS  {k}-PARTICIONES" if k == 2 else f"PRUEBAS  {k}-PARTICIONES"
        if k == 2:
            lbl = "PRUEBAS  BIPARTICIONES"
        ws.cell(3, cs, lbl).font = _bold(color="FFFFFF")
        ws.cell(3, cs).fill = _fill(COLOR_HDR_K)
        ws.merge_cells(start_row=3, start_column=cs, end_row=3, end_column=cs + 8)
        ws.cell(3, cs).alignment = Alignment(horizontal="center")

    # Fila 4: sub-headers de estrategia
    for k, cs in DST_START.items():
        for label, col_off, color in [
            ("QNodos",   0, COLOR_QNODOS),
            ("Geometric", 3, COLOR_GEO),
            ("Circuit",  6, COLOR_CIRCUIT),
        ]:
            c = cs + col_off
            ws.cell(4, c, label).font = _bold()
            ws.cell(4, c).fill = _fill(color)
            ws.merge_cells(start_row=4, start_column=c, end_row=4, end_column=c + 2)
            ws.cell(4, c).alignment = Alignment(horizontal="center")

    # Fila 5: etiquetas de columna
    for c, lbl in [(1, "#Prueba"), (2, "Alcance o Purview (t+1)"), (3, "Mecanismo(t)")]:
        ws.cell(5, c, lbl).font = _bold()

    for k, cs in DST_START.items():
        for offset, (color, etqs) in enumerate([
            (COLOR_QNODOS,  ["Partición", "Pérdida", "Tiempo"]),
            (COLOR_GEO,     ["Partición", "Pérdida", "Tiempo"]),
            (COLOR_CIRCUIT, ["Partición", "Pérdida", "Tiempo"]),
        ]):
            for i, lbl in enumerate(etqs):
                c = cs + offset * 3 + i
                cell = ws.cell(5, c, lbl)
                cell.font = _bold()
                cell.fill = _fill(color)
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Filas 6+: datos
    datos_src = list(ws_src.iter_rows(min_row=6, values_only=True))
    fila_dst = 6
    for fila in datos_src:
        if fila[1] is None:
            continue
        ws.cell(fila_dst, 2, fila[1])  # Alcance
        ws.cell(fila_dst, 3, fila[2])  # Mecanismo

        for k, cs in DST_START.items():
            q_cols, g_cols = ORIG_COLS[k]
            # QNodos (src 0-indexed → fila[col-1])
            for i, sc in enumerate(q_cols):
                ws.cell(fila_dst, cs + i, fila[sc - 1])
            # Geometric
            for i, sc in enumerate(g_cols):
                ws.cell(fila_dst, cs + 3 + i, fila[sc - 1])
            # Circuit: vacío (se rellena por run_circuito_single.py)

        fila_dst += 1

    # Anchos de columna
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    for k, cs in DST_START.items():
        for offset in range(9):
            col_ltr = get_column_letter(cs + offset)
            if offset in (0, 3, 6):   # columnas Partición
                ws.column_dimensions[col_ltr].width = 55
            else:
                ws.column_dimensions[col_ltr].width = 13

    print(f"  {tag}: {fila_dst - 6} filas")
    return fila_dst - 6


if __name__ == "__main__":
    wb_src = openpyxl.load_workbook(ORIGEN, data_only=True)
    wb_dst = openpyxl.Workbook()
    wb_dst.remove(wb_dst.active)

    print(f"Creando {DESTINO}...")
    for nombre_src, estado, sistema, tag in HOJAS:
        ws_src = wb_src[nombre_src]
        crear_hoja(wb_dst, ws_src, tag, estado, sistema)

    wb_dst.save(DESTINO)
    print(f"Guardado: {DESTINO}")
