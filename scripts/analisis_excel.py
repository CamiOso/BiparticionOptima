import openpyxl

wb = openpyxl.load_workbook('DatosPruebas2026_1.xlsx', data_only=True)

# ============================================================
# 1. 10A-Elementos: QNodos k=2 vs Geometric k=2
# ============================================================
ws = wb['10A-Elementos']
print('=== 10A-Elementos: QNodos k=2 vs Geometric k=2 ===')
acuerdos = 0
total = 0
q_tiempos = []
g_tiempos = []
for row in ws.iter_rows(min_row=8, max_row=57, values_only=True):
    if row[0] is None or not isinstance(row[0], (int, float)):
        continue
    total += 1
    q_perdida = row[4]
    g_perdida = row[7]
    q_tiempo = row[5]
    g_tiempo = row[8]
    if q_perdida is not None and g_perdida is not None:
        if abs(float(q_perdida) - float(g_perdida)) < 1e-6:
            acuerdos += 1
        q_tiempos.append(float(q_tiempo) if q_tiempo else 0)
        g_tiempos.append(float(g_tiempo) if g_tiempo else 0)

print(f'Total filas: {total}')
print(f'Acuerdos exactos: {acuerdos}/{total} ({100*acuerdos/total:.1f}%)')
if q_tiempos:
    print(f'Tiempo promedio QNodos: {sum(q_tiempos)/len(q_tiempos):.2f}s')
    print(f'Tiempo promedio Geometric: {sum(g_tiempos)/len(g_tiempos):.2f}s')
    print(f'Speedup Geometric/QNodos: {sum(q_tiempos)/sum(g_tiempos):.2f}x')

# ============================================================
# 2. 15B-Elementos: QNodos k=2 vs Geometric k=2
# ============================================================
ws2 = wb['15B-Elementos']
print()
print('=== 15B-Elementos: QNodos k=2 vs Geometric k=2 ===')
acuerdos2 = 0
total2 = 0
q_t2 = []
g_t2 = []
for row in ws2.iter_rows(min_row=8, max_row=57, values_only=True):
    if row[0] is None or not isinstance(row[0], (int, float)):
        continue
    total2 += 1
    q_perdida = row[4]
    g_perdida = row[7]
    q_tiempo = row[5]
    g_tiempo = row[8]
    if q_perdida is not None and g_perdida is not None:
        if abs(float(q_perdida) - float(g_perdida)) < 1e-6:
            acuerdos2 += 1
        q_t2.append(float(q_tiempo) if q_tiempo else 0)
        g_t2.append(float(g_tiempo) if g_tiempo else 0)

print(f'Total filas: {total2}')
print(f'Acuerdos exactos: {acuerdos2}/{total2} ({100*acuerdos2/total2:.1f}%)')
if q_t2:
    print(f'Tiempo promedio QNodos: {sum(q_t2)/len(q_t2):.2f}s')
    print(f'Tiempo promedio Geometric: {sum(g_t2)/len(g_t2):.2f}s')
    print(f'Speedup Geometric/QNodos: {sum(q_t2)/sum(g_t2):.2f}x')

# ============================================================
# 3. 22A-Elementos: k=2 vs k=3,4,5 (QNodos)
# ============================================================
ws3 = wb['22A-Elementos']
print()
print('=== 22A-Elementos: k=2 vs k=3,4,5 (QNodos) ===')
k2_vs_k3 = []
for row in ws3.iter_rows(min_row=8, max_row=57, values_only=True):
    if row[0] is None or not isinstance(row[0], (int, float)):
        continue
    q_k2 = row[4]
    q_k3 = row[10]
    q_k4 = row[16]
    q_k5 = row[22]
    if q_k2 and q_k3:
        k2_vs_k3.append((float(q_k2), float(q_k3), float(q_k4) if q_k4 else 0, float(q_k5) if q_k5 else 0))

print(f'Filas con datos k=2 y k=3: {len(k2_vs_k3)}')
for i, (k2, k3, k4, k5) in enumerate(k2_vs_k3[:10]):
    print(f'  Fila {i}: k2={k2:.6f}  k3={k3:.6f}  k4={k4:.6f}  k5={k5:.6f}  ratio k3/k2={k3/k2:.1f}x')

k2_min = all(k2 <= k3 and k2 <= k4 and k2 <= k5 for k2, k3, k4, k5 in k2_vs_k3)
print(f'k=2 es siempre el mínimo: {k2_min}')

# ============================================================
# 4. 25A-Elementos: k=2 vs k=3,4,5 (QNodos)
# ============================================================
ws4 = wb['25A-Elementos']
print()
print('=== 25A-Elementos: k=2 vs k=3,4,5 (QNodos) ===')
k2_vs_k3_25 = []
for row in ws4.iter_rows(min_row=8, max_row=57, values_only=True):
    if row[0] is None or not isinstance(row[0], (int, float)):
        continue
    q_k2 = row[4]
    q_k3 = row[10]
    q_k4 = row[16]
    q_k5 = row[22]
    if q_k2 and q_k3:
        k2_vs_k3_25.append((float(q_k2), float(q_k3), float(q_k4) if q_k4 else 0, float(q_k5) if q_k5 else 0))

print(f'Filas con datos k=2 y k=3: {len(k2_vs_k3_25)}')
for i, (k2, k3, k4, k5) in enumerate(k2_vs_k3_25[:10]):
    print(f'  Fila {i}: k2={k2:.6f}  k3={k3:.6f}  k4={k4:.6f}  k5={k5:.6f}  ratio k3/k2={k3/k2:.1f}x')

k2_min_25 = all(k2 <= k3 and k2 <= k4 and k2 <= k5 for k2, k3, k4, k5 in k2_vs_k3_25)
print(f'k=2 es siempre el mínimo: {k2_min_25}')

# ============================================================
# 5. Resumen de tiempos por k
# ============================================================
print()
print('=== Resumen tiempos QNodos por k (22A) ===')
for row in ws3.iter_rows(min_row=8, max_row=57, values_only=True):
    if row[0] is None or not isinstance(row[0], (int, float)):
        continue
    t_k2 = row[5]
    t_k3 = row[11]
    t_k4 = row[17]
    t_k5 = row[23]
    if t_k2 and t_k3:
        print(f'  t_k2={float(t_k2):.2f}s  t_k3={float(t_k3):.2f}s  t_k4={float(t_k4):.2f}s  t_k5={float(t_k5):.2f}s')
        break  # solo primera fila como muestra

print()
print('=== Resumen tiempos QNodos por k (25A) ===')
for row in ws4.iter_rows(min_row=8, max_row=57, values_only=True):
    if row[0] is None or not isinstance(row[0], (int, float)):
        continue
    t_k2 = row[5]
    t_k3 = row[11]
    t_k4 = row[17]
    t_k5 = row[23]
    if t_k2 and t_k3:
        print(f'  t_k2={float(t_k2):.2f}s  t_k3={float(t_k3):.2f}s  t_k4={float(t_k4):.2f}s  t_k5={float(t_k5):.2f}s')
        break
