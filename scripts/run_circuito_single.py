"""Circuito para UNA fila. Usa DatosCircuito2026.xlsx.

Uso:
  python scripts/run_circuito_single.py 6                  # 20A por defecto
  python scripts/run_circuito_single.py 6 --sistema 10A
  python scripts/run_circuito_single.py 6 --sistema 25A
  python scripts/run_circuito_single.py 6 --sistema 10A --start-k 3
"""
import sys, time, gc, argparse
import numpy as np
import openpyxl

PROJECT = "/home/cami/Desktop/AnalisisDiseñoAlgoritmos/Proyecto/ProyectoAnalisis2026"
sys.path.insert(0, PROJECT)

SISTEMAS = {
    "10A": {
        "excel": "DatosCircuito2026.xlsx",
        "sheet": "10A",
        "tpm": "src/.samples/N10A.csv",
        "tpm_fmt": "csv",
        "sistema": "ABCDEFGHIJ",
        "estado": "1000000000",
        "n": 10,
    },
    "15B": {
        "excel": "DatosCircuito2026.xlsx",
        "sheet": "15B",
        "tpm": "src/.samples/N15B.csv",
        "tpm_fmt": "csv",
        "sistema": "ABCDEFGHIJKLMNO",
        "estado": "100000000000000",
        "n": 15,
    },
    "20A": {
        "excel": "DatosCircuito2026.xlsx",
        "sheet": "20A",
        "tpm": "src/.samples/N20A.npy",
        "tpm_fmt": "npy",
        "sistema": "ABCDEFGHIJKLMNOPQRST",
        "estado": "10000000000000000000",
        "n": 20,
    },
    "22A": {
        "excel": "DatosCircuito2026.xlsx",
        "sheet": "22A",
        "tpm": "src/.samples/N22A.npy",
        "tpm_fmt": "npy",
        "sistema": "ABCDEFGHIJKLMNOPQRSTUV",
        "estado": "1000000000000000000000",
        "n": 22,
    },
    "25A": {
        "excel": "DatosCircuito2026.xlsx",
        "sheet": "25A",
        "tpm": "src/.samples/N25A.npy",
        "tpm_fmt": "memmap",
        "sistema": "ABCDEFGHIJKLMNOPQRSTUVWXY",
        "estado": "1000000000000000000000000",
        "n": 25,
    },
}

# Columnas Circuit en DatosCircuito2026.xlsx (Partición, Pérdida, Tiempo)
CIRCUIT_COLS = {2: (10, 11, 12), 3: (19, 20, 21), 4: (28, 29, 30), 5: (37, 38, 39)}


def to_mask(sub, sistema):
    return "".join("1" if c in sub else "0" for c in sistema)


def guardar_k(excel_path, sheet, fila_idx, k, part, perd, elapsed):
    import os, random, time as _t
    cp, cl, ct = CIRCUIT_COLS[k]
    lock = excel_path + ".lock"
    for _ in range(20):
        try:
            fd = os.open(lock, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.close(fd)
            break
        except FileExistsError:
            _t.sleep(random.uniform(0.5, 2.0))
    else:
        raise RuntimeError("No se pudo adquirir lock del Excel")
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb[sheet]
        ws.cell(row=fila_idx, column=cp, value=part)
        ws.cell(row=fila_idx, column=cl, value=perd)
        ws.cell(row=fila_idx, column=ct, value=round(elapsed, 4))
        wb.save(excel_path)
        wb.close()
    finally:
        try:
            os.unlink(lock)
        except FileNotFoundError:
            pass


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("fila", type=int)
    parser.add_argument("--sistema", type=str, default="20A",
                        choices=list(SISTEMAS.keys()))
    parser.add_argument("--start-k", type=int, default=2, choices=[2, 3, 4, 5])
    parser.add_argument("--end-k", type=int, default=5, choices=[2, 3, 4, 5])
    args = parser.parse_args()

    cfg = SISTEMAS[args.sistema]
    excel_path = f"{PROJECT}/{cfg['excel']}"
    tpm_path = f"{PROJECT}/{cfg['tpm']}"
    sheet = cfg["sheet"]
    sistema_str = cfg["sistema"]
    estado = cfg["estado"]
    n = cfg["n"]
    condicion = "1" * n
    fila_idx = args.fila

    # Cache según n del sistema
    if n <= 13:
        _cn, _cs = 512, 1024
    elif n <= 17:
        _cn, _cs = 128, 512
    elif n <= 20:
        _cn, _cs = 64, 256
    else:
        _cn, _cs = 32, 128
    import src.modelos.nucleo.ncubo as _ncubo_mod
    import src.modelos.nucleo.sistema as _sistema_mod
    _ncubo_mod._MAX_MEMO_NCUBE = _cn
    _sistema_mod._MAX_MEMO_SISTEMA = _cs

    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet]
    alc_letras = str(ws.cell(fila_idx, 2).value or "").strip()
    mec_letras = str(ws.cell(fila_idx, 3).value or "").strip()
    wb.close()

    if not alc_letras:
        print(f"[fila={fila_idx}] alcance vacio, abortando", flush=True)
        sys.exit(0)

    alc_mask = to_mask(alc_letras, sistema_str)
    mec_mask = to_mask(mec_letras, sistema_str)
    n_max = max(len(alc_letras), len(mec_letras))
    print(f"[{args.sistema} fila={fila_idx}] n_max={n_max} | alc={alc_letras} mec={mec_letras} | cache ncube={_cn} sistema={_cs}", flush=True)

    from src.estrategias.circuito import Circuito
    print(f"[{args.sistema} fila={fila_idx}] Cargando TPM...", flush=True)
    fmt = cfg["tpm_fmt"]
    if fmt == "csv":
        tpm = np.loadtxt(tpm_path, delimiter=",", dtype=np.float64)
    elif fmt == "memmap":
        tpm = np.memmap(tpm_path, dtype=np.float32, mode="r", shape=(2**n, n))
    else:
        tpm = np.load(tpm_path, mmap_mode="r")
    circuito = Circuito(tpm)

    t_fila = time.perf_counter()
    for k in (k for k in (2, 3, 4, 5) if args.start_k <= k <= args.end_k):
        t0 = time.perf_counter()
        res = circuito.aplicar_estrategia(
            estado_inicial=estado, condicion=condicion,
            alcance=alc_mask, mecanismo=mec_mask, k=k,
        )
        elapsed = time.perf_counter() - t0
        part = str(res.particion) if hasattr(res, "particion") else ""
        perd = float(res.perdida)
        print(f"  [{args.sistema} fila={fila_idx} k={k}] perdida={round(perd,6)} t={round(elapsed,2)}s", flush=True)
        guardar_k(excel_path, sheet, fila_idx, k, part, perd, elapsed)
        print(f"  [{args.sistema} fila={fila_idx} k={k}] GUARDADO", flush=True)
        gc.collect()

    total = round(time.perf_counter() - t_fila, 1)
    print(f"  [{args.sistema} fila={fila_idx}] COMPLETO {total}s", flush=True)
