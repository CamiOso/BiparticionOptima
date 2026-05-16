"""QNodos + Geometric paralelo (4 núcleos) para 9 casos pequeños de 22A-Elementos.
Usa fork de Linux: TPM cargada en proceso padre, workers la heredan sin pickle.
"""
import sys, time, gc
import numpy as np
import openpyxl

PROJECT = "/home/cami/Desktop/AnalisisDiseñoAlgoritmos/Proyecto/ProyectoAnalisis2026"
sys.path.insert(0, PROJECT)

EXCEL     = f"{PROJECT}/DatosPruebas2026_1.xlsx"
SHEET     = "22A-Elementos"
CSV       = f"{PROJECT}/src/.samples/N22A.npy"
SISTEMA   = "ABCDEFGHIJKLMNOPQRSTUV"
ESTADO    = "1000000000000000000000"
CONDICION = "1" * 22

# Columnas 1-indexed (misma estructura que 20A)
QNODOS_COLS  = {2: (4,5,6),   3: (10,11,12), 4: (16,17,18), 5: (22,23,24)}
GEO_COLS     = {2: (7,8,9),   3: (13,14,15), 4: (19,20,21), 5: (25,26,27)}

FILAS = [
    (38, "ABDEGHJKMNPQSTV", "ABDEGHJKMNPQSTV"),
    (39, "ABDEGHJKMNPQSTV", "ACEGIKMOQSU"),
    (40, "ABDEGHJKMNPQSTV", "BDFHJLNPRTV"),
    (45, "ACEGIKMOQSU",     "ABDEGHJKMNPQSTV"),
    (46, "ACEGIKMOQSU",     "ACEGIKMOQSU"),
    (47, "ACEGIKMOQSU",     "BDFHJLNPRTV"),
    (52, "BDFHJLNPRTV",     "ABDEGHJKMNPQSTV"),
    (53, "BDFHJLNPRTV",     "ACEGIKMOQSU"),
    (54, "BDFHJLNPRTV",     "BDFHJLNPRTV"),
]

# TPM global — heredada por workers via fork (sin pickle)
_tpm = None

def to_mask(sub, sistema):
    return "".join("1" if c in sub else "0" for c in sistema)

def procesar_fila(args):
    from src.estrategias.q_nodos import QNodos
    from src.strategies.geometric import Geometric
    fila_idx, alc_letras, mec_letras = args
    alc_mask = to_mask(alc_letras, SISTEMA)
    mec_mask = to_mask(mec_letras, SISTEMA)
    n_max = max(len(alc_letras), len(mec_letras))
    qnodos = QNodos(_tpm)
    geo    = Geometric(_tpm)
    geo._usar_paralelizacion_costos = False  # evitar hilos internos — ya somos un worker
    resultado = {"qnodos": {}, "geometric": {}}
    t_fila = time.perf_counter()
    for k in (2, 3, 4, 5):
        for nombre, estrategia, destino in [("QNodos", qnodos, "qnodos"),
                                            ("Geometric", geo, "geometric")]:
            t0 = time.perf_counter()
            res = estrategia.aplicar_estrategia(
                estado_inicial=ESTADO,
                condicion=CONDICION,
                alcance=alc_mask,
                mecanismo=mec_mask,
                k=k,
            )
            elapsed = time.perf_counter() - t0
            resultado[destino][k] = (
                str(res.particion) if hasattr(res,"particion") else "",
                float(res.perdida), round(elapsed, 4)
            )
            print(f"  [fila={fila_idx} k={k} {nombre}] perdida={round(float(res.perdida),6)} t={round(elapsed,2)}s", flush=True)
            gc.collect()
    total = round(time.perf_counter() - t_fila, 1)
    print(f"  [fila={fila_idx}] LISTO {total}s  n_max={n_max}", flush=True)
    return fila_idx, resultado

def guardar(fila_idx, resultado):
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[SHEET]
    for k, (part, perd, t) in resultado["qnodos"].items():
        cp, cl, ct = QNODOS_COLS[k]
        ws.cell(row=fila_idx, column=cp, value=part)
        ws.cell(row=fila_idx, column=cl, value=perd)
        ws.cell(row=fila_idx, column=ct, value=t)
    for k, (part, perd, t) in resultado["geometric"].items():
        cp, cl, ct = GEO_COLS[k]
        ws.cell(row=fila_idx, column=cp, value=part)
        ws.cell(row=fila_idx, column=cl, value=perd)
        ws.cell(row=fila_idx, column=ct, value=t)
    wb.save(EXCEL)
    wb.close()
    print(f"[GUARDADO] fila {fila_idx} (QNodos + Geometric)", flush=True)

if __name__ == "__main__":
    import os, multiprocessing as mp
    os.environ["OMP_NUM_THREADS"] = "1"
    os.environ["OPENBLAS_NUM_THREADS"] = "1"
    os.environ["MKL_NUM_THREADS"] = "1"
    mp.set_start_method("fork", force=True)

    print("Cargando TPM 22A (float32)...", flush=True)
    globals()["_tpm"] = np.load(CSV)
    print(f"TPM: {_tpm.shape}  | Lanzando Pool(4) — fork, sin pickle...", flush=True)
    t0 = time.perf_counter()
    with mp.Pool(4) as pool:
        for fila_idx, resultado in pool.imap_unordered(procesar_fila, FILAS):
            guardar(fila_idx, resultado)
    print(f"\nTodas las filas completadas en {round((time.perf_counter()-t0)/60,1)} min", flush=True)
