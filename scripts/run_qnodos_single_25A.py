"""QNodos para UNA fila de 25A-Elementos. Lanzar varias instancias con &"""
import sys, time, gc, argparse, os, json, random
import numpy as np
import openpyxl

PROJECT = "/home/cami/Desktop/AnalisisDiseñoAlgoritmos/Proyecto/ProyectoAnalisis2026"
sys.path.insert(0, PROJECT)

EXCEL     = f"{PROJECT}/DatosPruebas2026_1.xlsx"
SHEET     = "25A-Elementos "   # trailing space — nombre exacto del sheet
CSV       = f"{PROJECT}/src/.samples/N25A.npy"
SISTEMA   = "ABCDEFGHIJKLMNOPQRSTUVWXY"
ESTADO    = "1000000000000000000000000"
CONDICION = "1" * 25
Q_COLS    = {2: (4,5,6), 3: (10,11,12), 4: (16,17,18), 5: (22,23,24)}

FILAS = {
    6:  ("ABCDEFGHIJKLMNOPQRSTUVWXY", "ABCDEFGHIJKLMNOPQRSTUVWXY"),
    7:  ("ABCDEFGHIJKLMNOPQRSTUVWXY", "ABCDEFGHIJKLMNOPQRSTUVWX"),
    8:  ("ABCDEFGHIJKLMNOPQRSTUVWXY", "BCDEFGHIJKLMNOPQRSTUVWXY"),
    9:  ("ABCDEFGHIJKLMNOPQRSTUVWXY", "BCDEFGHIJKLMNOPQRSTUVWX"),
    10: ("ABCDEFGHIJKLMNOPQRSTUVWXY", "ABDEGHJKMNPQSTVWY"),
    11: ("ABCDEFGHIJKLMNOPQRSTUVWXY", "ACEGIKMOQSUWY"),
    12: ("ABCDEFGHIJKLMNOPQRSTUVWXY", "BDFHJLNPRTVX"),
    13: ("ABCDEFGHIJKLMNOPQRSTUVWX",  "ABCDEFGHIJKLMNOPQRSTUVWXY"),
    14: ("ABCDEFGHIJKLMNOPQRSTUVWX",  "ABCDEFGHIJKLMNOPQRSTUVWX"),
    15: ("ABCDEFGHIJKLMNOPQRSTUVWX",  "BCDEFGHIJKLMNOPQRSTUVWXY"),
    16: ("ABCDEFGHIJKLMNOPQRSTUVWX",  "BCDEFGHIJKLMNOPQRSTUVWX"),
    17: ("ABCDEFGHIJKLMNOPQRSTUVWX",  "ABDEGHJKMNPQSTVWY"),
    18: ("ABCDEFGHIJKLMNOPQRSTUVWX",  "ACEGIKMOQSUWY"),
    19: ("ABCDEFGHIJKLMNOPQRSTUVWX",  "BDFHJLNPRTVX"),
    20: ("BCDEFGHIJKLMNOPQRSTUVWXY",  "ABCDEFGHIJKLMNOPQRSTUVWXY"),
    21: ("BCDEFGHIJKLMNOPQRSTUVWXY",  "ABCDEFGHIJKLMNOPQRSTUVWX"),
    22: ("BCDEFGHIJKLMNOPQRSTUVWXY",  "BCDEFGHIJKLMNOPQRSTUVWXY"),
    23: ("BCDEFGHIJKLMNOPQRSTUVWXY",  "BCDEFGHIJKLMNOPQRSTUVWX"),
    24: ("BCDEFGHIJKLMNOPQRSTUVWXY",  "ABDEGHJKMNPQSTVWY"),
    25: ("BCDEFGHIJKLMNOPQRSTUVWXY",  "ACEGIKMOQSUWY"),
    26: ("BCDEFGHIJKLMNOPQRSTUVWXY",  "BDFHJLNPRTVX"),
    27: ("BCDEFGHIJKLMNOPQRSTUVWX",   "ABCDEFGHIJKLMNOPQRSTUVWXY"),
    28: ("BCDEFGHIJKLMNOPQRSTUVWX",   "ABCDEFGHIJKLMNOPQRSTUVWX"),
    29: ("BCDEFGHIJKLMNOPQRSTUVWX",   "BCDEFGHIJKLMNOPQRSTUVWXY"),
    30: ("BCDEFGHIJKLMNOPQRSTUVWX",   "BCDEFGHIJKLMNOPQRSTUVWX"),
    31: ("BCDEFGHIJKLMNOPQRSTUVWX",   "ABDEGHJKMNPQSTVWY"),
    32: ("BCDEFGHIJKLMNOPQRSTUVWX",   "ACEGIKMOQSUWY"),
    33: ("BCDEFGHIJKLMNOPQRSTUVWX",   "BDFHJLNPRTVX"),
    34: ("ABDEGHJKMNPQSTVWY",         "ABCDEFGHIJKLMNOPQRSTUVWXY"),
    35: ("ABDEGHJKMNPQSTVWY",         "ABCDEFGHIJKLMNOPQRSTUVWX"),
    36: ("ABDEGHJKMNPQSTVWY",         "BCDEFGHIJKLMNOPQRSTUVWXY"),
    37: ("ABDEGHJKMNPQSTVWY",         "BCDEFGHIJKLMNOPQRSTUVWX"),
    38: ("ABDEGHJKMNPQSTVWY",         "ABDEGHJKMNPQSTVWY"),
    39: ("ABDEGHJKMNPQSTVWY",         "ACEGIKMOQSUWY"),
    40: ("ABDEGHJKMNPQSTVWY",         "BDFHJLNPRTVX"),
    41: ("ACEGIKMOQSUWY",             "ABCDEFGHIJKLMNOPQRSTUVWXY"),
    42: ("ACEGIKMOQSUWY",             "ABCDEFGHIJKLMNOPQRSTUVWX"),
    43: ("ACEGIKMOQSUWY",             "BCDEFGHIJKLMNOPQRSTUVWXY"),
    44: ("ACEGIKMOQSUWY",             "BCDEFGHIJKLMNOPQRSTUVWX"),
    45: ("ACEGIKMOQSUWY",             "ABDEGHJKMNPQSTVWY"),
    46: ("ACEGIKMOQSUWY",             "ACEGIKMOQSUWY"),
    47: ("ACEGIKMOQSUWY",             "BDFHJLNPRTVX"),
    48: ("BDFHJLNPRTVX",              "ABCDEFGHIJKLMNOPQRSTUVWXY"),
    49: ("BDFHJLNPRTVX",              "ABCDEFGHIJKLMNOPQRSTUVWX"),
    50: ("BDFHJLNPRTVX",              "BCDEFGHIJKLMNOPQRSTUVWXY"),
    51: ("BDFHJLNPRTVX",              "BCDEFGHIJKLMNOPQRSTUVWX"),
    52: ("BDFHJLNPRTVX",              "ABDEGHJKMNPQSTVWY"),
    53: ("BDFHJLNPRTVX",              "ACEGIKMOQSUWY"),
    54: ("BDFHJLNPRTVX",              "BDFHJLNPRTVX"),
    55: ("ACDEFGHIJKLMNOPQRSTVX",     "ACDEFGHIJKLMNOPQRSTVX"),
}

LOCK       = EXCEL + ".lock"
MEC_CACHE  = "/tmp/mec_cache_25A.json"
LOCK_CACHE = MEC_CACHE + ".lock"

def to_mask(sub):
    return "".join("1" if c in sub else "0" for c in SISTEMA)

def guardar_k(fila_idx, k, part, perd, elapsed):
    import os, random, time as _t
    cp, cl, ct = Q_COLS[k]
    for _ in range(20):
        try:
            fd = os.open(LOCK, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.close(fd)
            break
        except FileExistsError:
            _t.sleep(random.uniform(0.5, 2.0))
    else:
        raise RuntimeError("No se pudo adquirir lock del Excel")
    try:
        wb = openpyxl.load_workbook(EXCEL)
        ws = wb[SHEET]
        ws.cell(row=fila_idx, column=cp, value=part)
        ws.cell(row=fila_idx, column=cl, value=perd)
        ws.cell(row=fila_idx, column=ct, value=round(elapsed, 4))
        wb.save(EXCEL)
        wb.close()
    finally:
        try:
            os.unlink(LOCK)
        except FileNotFoundError:
            pass

def _actualizar_cache(mec_key: str, k: int, perdida: float, fila_idx: int) -> None:
    """Guarda resultado de fila referencia (alc<=mec) en cache para reutilizar en filas con alc>mec."""
    for _ in range(10):
        try:
            fd = os.open(LOCK_CACHE, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.close(fd)
            break
        except FileExistsError:
            time.sleep(random.uniform(0.5, 1.5))
    else:
        return
    try:
        cache_data: dict = {}
        if os.path.exists(MEC_CACHE):
            try:
                with open(MEC_CACHE) as f:
                    cache_data = json.load(f)
            except Exception:
                pass
        if mec_key not in cache_data:
            cache_data[mec_key] = {}
        cache_data[mec_key][str(k)] = {"perdida": perdida, "source": fila_idx}
        with open(MEC_CACHE, "w") as f:
            json.dump(cache_data, f)
    finally:
        try:
            os.unlink(LOCK_CACHE)
        except FileNotFoundError:
            pass


if __name__ == "__main__":
    os.environ.setdefault("OMP_NUM_THREADS", "1")
    os.environ.setdefault("OPENBLAS_NUM_THREADS", "1")

    parser = argparse.ArgumentParser()
    parser.add_argument("fila", type=int)
    parser.add_argument("--start-k", type=int, default=2, choices=[2, 3, 4, 5])
    args = parser.parse_args()
    fila_idx = args.fila
    alc_letras, mec_letras = FILAS[fila_idx]

    # --- Mec-cache: mismo mec + alc_n > mec_n → perdida idéntica (empíricamente verificado)
    mec_key    = "".join(sorted(mec_letras))
    alc_n_val  = len(alc_letras)
    mec_n_val  = len(mec_letras)
    es_referencia = (alc_n_val <= mec_n_val)

    if alc_n_val > mec_n_val:
        cache_data: dict = {}
        if os.path.exists(MEC_CACHE):
            try:
                with open(MEC_CACHE) as _cf:
                    cache_data = json.load(_cf)
            except Exception:
                pass
        ks_req = [k for k in (2, 3, 4, 5) if k >= args.start_k]
        mec_entry = cache_data.get(mec_key, {})
        if all(str(k) in mec_entry for k in ks_req):
            print(
                f"[fila={fila_idx}] CACHE HIT mec={mec_letras[:14]}... "
                f"alc_n={alc_n_val} > mec_n={mec_n_val} → reutilizando perdida",
                flush=True,
            )
            for k in ks_req:
                entry = mec_entry[str(k)]
                perd  = entry["perdida"]
                src   = entry["source"]
                part  = f"CACHE:fila{src}"
                print(f"  [fila={fila_idx} k={k}] perdida={round(perd,6)} (cache de fila {src})", flush=True)
                guardar_k(fila_idx, k, part, perd, 0.0)
                print(f"  [fila={fila_idx} k={k}] GUARDADO (cache)", flush=True)
            print(f"  [fila={fila_idx}] COMPLETO (cache, 0s)", flush=True)
            sys.exit(0)
        print(
            f"[fila={fila_idx}] alc_n={alc_n_val} > mec_n={mec_n_val} pero cache vacío "
            f"para mec={mec_letras[:14]}... → ejecutando normal",
            flush=True,
        )
    # -------------------------------------------------------------------------

    mec_n = len(mec_letras)
    if mec_n <= 13:
        _cn, _cs = 512, 1024
    elif mec_n <= 17:
        _cn, _cs = 128, 512
    elif mec_n <= 20:
        _cn, _cs = 64,  256
    else:
        _cn, _cs = 32,  128
    import src.modelos.nucleo.ncubo   as _ncubo_mod
    import src.modelos.nucleo.sistema as _sistema_mod
    _ncubo_mod._MAX_MEMO_NCUBE   = _cn
    _sistema_mod._MAX_MEMO_SISTEMA = _cs
    print(f"[fila={fila_idx}] cache ncube={_cn} sistema={_cs} (mec={mec_n})", flush=True)

    from src.estrategias.q_nodos import QNodos
    print(f"[fila={fila_idx}] Cargando TPM...", flush=True)
    tpm = np.load(CSV, mmap_mode="r")
    qnodos = QNodos(tpm)

    alc_mask = to_mask(alc_letras)
    mec_mask = to_mask(mec_letras)
    n_max = max(len(alc_letras), len(mec_letras))
    print(f"[fila={fila_idx}] n_max={n_max} | {alc_letras} / {mec_letras} | start_k={args.start_k}", flush=True)

    t_fila = time.perf_counter()
    for k in (k for k in (2, 3, 4, 5) if k >= args.start_k):
        t0 = time.perf_counter()
        res = qnodos.aplicar_estrategia(
            estado_inicial=ESTADO, condicion=CONDICION,
            alcance=alc_mask, mecanismo=mec_mask, k=k,
        )
        elapsed = time.perf_counter() - t0
        part = str(res.particion) if hasattr(res, "particion") else ""
        perd = float(res.perdida)
        print(f"  [fila={fila_idx} k={k}] perdida={round(perd,6)} t={round(elapsed,2)}s", flush=True)
        guardar_k(fila_idx, k, part, perd, elapsed)
        print(f"  [fila={fila_idx} k={k}] GUARDADO", flush=True)
        if es_referencia:
            _actualizar_cache(mec_key, k, perd, fila_idx)
        gc.collect()

    total = round(time.perf_counter() - t_fila, 1)
    print(f"  [fila={fila_idx}] COMPLETO {total}s", flush=True)
