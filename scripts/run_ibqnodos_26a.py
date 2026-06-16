"""Benchmark IBQNodos en hoja 26A-Elementos de DatosIBQNodos2026.xlsx.

Lee los casos de la hoja 26A-Elementos (columnas B=Alcance, C=Mecanismo),
ejecuta IBQNodos con warm-start y escribe los resultados en las columnas
D=Partición, E=Pérdida, F=Tiempo de la misma hoja.

Uso:
    python run_ibqnodos_26a.py
    python run_ibqnodos_26a.py --desde 10   # retomar desde fila Excel 10
"""

import argparse
import json
import re
import signal
import sys
import time
from pathlib import Path

import numpy as np
import openpyxl

PROJECT = Path(__file__).parent.parent.resolve()
sys.path.insert(0, str(PROJECT))

EXCEL       = PROJECT / "DatosIBQNodos2026.xlsx"
CHECKPOINT  = PROJECT / "checkpoint_ibqnodos_26a.json"
TPM_PATH    = PROJECT / "src/.samples/N26A.npy"

N       = 26
SISTEMA = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
ESTADO  = "1" + "0" * 25
TIMEOUT_CASO = 3600   # 1 hora por caso


class _CasoTimeout(Exception):
    pass


def _sigalrm_handler(signum, frame):
    raise _CasoTimeout()


# ── Helpers ────────────────────────────────────────────────────────────────

def to_mask(letters: str) -> str:
    s = set(letters.upper())
    return "".join("1" if c in s else "0" for c in SISTEMA)


_RE_GRUPO_A = re.compile(r'\(M=\(([^)]*)\), A=\(([^)]*)\)\)')

def _parse_grupo_a(particion_str: str) -> "set | None":
    m = _RE_GRUPO_A.match(particion_str or "")
    if not m:
        return None
    def nums(s: str) -> list[int]:
        return [int(x.strip()) for x in s.split(",") if x.strip().lstrip("-").isdigit()]
    grupo_a: set = set()
    for mn in nums(m.group(1)):
        grupo_a.add((0, mn))
    for an in nums(m.group(2)):
        grupo_a.add((1, an))
    return grupo_a or None


# ── Checkpoint ────────────────────────────────────────────────────────────

def guardar_checkpoint(data: dict) -> None:
    CHECKPOINT.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def cargar_checkpoint() -> dict:
    if CHECKPOINT.exists():
        return json.loads(CHECKPOINT.read_text(encoding="utf-8"))
    return {}


# ── Leer casos de Excel ───────────────────────────────────────────────────

def leer_casos() -> list[dict]:
    wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)
    ws = wb["26A-Elementos"]
    casos = []
    for r in range(6, ws.max_row + 1):
        alc = ws.cell(r, 2).value
        mec = ws.cell(r, 3).value
        if alc is None:
            break
        casos.append({
            "fila":    r,
            "alc_str": str(alc),
            "mec_str": str(mec),
            "alc_bin": to_mask(str(alc)),
            "mec_bin": to_mask(str(mec)),
        })
    wb.close()
    return casos


# ── Hoja resumen IBQNodos ─────────────────────────────────────────────────

HOJA_IB  = "26A-IBQNodos"
HDRS_IB  = ["#Prueba", "Alcance", "Mecanismo", "IBQNodos_φ", "IBQNodos_t(s)", "Partición_IBQNodos"]

def _inicializar_hoja_ibqnodos() -> None:
    wb = openpyxl.load_workbook(EXCEL)
    if HOJA_IB not in wb.sheetnames:
        ws = wb.create_sheet(HOJA_IB)
    else:
        ws = wb[HOJA_IB]
    if ws.cell(1, 1).value != HDRS_IB[0]:
        for col, h in enumerate(HDRS_IB, 1):
            ws.cell(1, col, h)
    wb.save(EXCEL)
    wb.close()
    print(f"Hoja '{HOJA_IB}' lista.", flush=True)


# ── Escribir resultado en Excel ───────────────────────────────────────────

def escribir_resultado_excel(fila: int, particion: str, perdida: float | None, t: float | None,
                              prueba: int, alc_str: str, mec_str: str) -> None:
    wb = openpyxl.load_workbook(EXCEL)
    # 26A-Elementos
    ws = wb["26A-Elementos"]
    ws.cell(fila, 4, particion)
    ws.cell(fila, 5, round(perdida, 8) if perdida is not None else "TIMEOUT")
    ws.cell(fila, 6, round(t, 3)       if t       is not None else None)
    # 26A-IBQNodos
    ws2 = wb[HOJA_IB]
    fila2 = prueba + 1   # fila 1 = encabezados
    ws2.cell(fila2, 1, prueba)
    ws2.cell(fila2, 2, alc_str)
    ws2.cell(fila2, 3, mec_str)
    ws2.cell(fila2, 4, round(perdida, 8) if perdida is not None else "TIMEOUT")
    ws2.cell(fila2, 5, round(t, 3)       if t       is not None else None)
    ws2.cell(fila2, 6, particion)
    wb.save(EXCEL)
    wb.close()


# ── Main ──────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--desde", type=int, default=0,
                        help="Número de fila Excel desde la que empezar.")
    args = parser.parse_args()

    if not TPM_PATH.exists():
        print(f"ERROR: {TPM_PATH} no existe. Ejecuta: python generar_N26A.py", file=sys.stderr)
        sys.exit(1)

    print(f"Cargando TPM {TPM_PATH.name} ({TPM_PATH.stat().st_size/1e9:.2f} GB) ...", flush=True)
    t0 = time.perf_counter()
    tpm = np.load(TPM_PATH, mmap_mode="r")
    print(f"TPM cargada en {time.perf_counter()-t0:.1f}s | shape={tpm.shape}", flush=True)

    from src.estrategias.ib_qnodos import IBQNodos
    estrategia = IBQNodos(tpm)

    _inicializar_hoja_ibqnodos()
    casos      = leer_casos()
    checkpoint = cargar_checkpoint()
    done_filas = {c["fila"] for c in checkpoint.get("casos", []) if "phi_ib" in c}

    # Seed cache desde checkpoint
    seed_cache: dict[str, set] = {}
    for c in checkpoint.get("casos", []):
        if c.get("phi_ib") is not None:
            ga = _parse_grupo_a(c.get("particion_ib", ""))
            if ga:
                seed_cache[c["mec_bin"]] = ga

    condicion = "1" * N

    print(f"\n{'='*60}")
    print(f"  26A: {len(casos)} casos  |  completados: {len(done_filas)}")
    print(f"{'='*60}\n", flush=True)

    resultados = list(checkpoint.get("casos", []))
    t_total = time.perf_counter()

    for i, caso in enumerate(casos):
        if caso["fila"] < args.desde:
            continue
        if caso["fila"] in done_filas:
            print(f"  [{i+1}/{len(casos)}] fila={caso['fila']} — ya completado, saltando.", flush=True)
            continue

        n_alc = caso["alc_bin"].count("1")
        n_mec = caso["mec_bin"].count("1")
        print(f"  [{i+1}/{len(casos)}] fila={caso['fila']} "
              f"alc={caso['alc_str'][:10]}...({n_alc}) mec={caso['mec_str'][:10]}...({n_mec})", flush=True)

        grupo_a_seed = seed_cache.get(caso["mec_bin"])
        if grupo_a_seed:
            print(f"    warm-start: seed_size={len(grupo_a_seed)}", flush=True)

        try:
            signal.signal(signal.SIGALRM, _sigalrm_handler)
            signal.alarm(TIMEOUT_CASO)
            t0 = time.perf_counter()
            try:
                res = estrategia.aplicar_estrategia(
                    estado_inicial=ESTADO,
                    condicion=condicion,
                    alcance=caso["alc_bin"],
                    mecanismo=caso["mec_bin"],
                    k=2,
                    grupo_a_seed=grupo_a_seed,
                )
            finally:
                signal.alarm(0)
            elapsed = time.perf_counter() - t0

            phi_ib = float(res.perdida)
            print(f"    φ={phi_ib:.6f}  t={elapsed:.1f}s", flush=True)

            caso_resultado = {**caso, "phi_ib": phi_ib, "t_ib": round(elapsed, 3),
                              "particion_ib": str(res.particion)}
            ga_new = _parse_grupo_a(str(res.particion))
            if ga_new:
                seed_cache[caso["mec_bin"]] = ga_new

            prueba = caso["fila"] - 5
            escribir_resultado_excel(caso["fila"], str(res.particion), phi_ib, elapsed,
                                     prueba, caso["alc_str"], caso["mec_str"])

        except _CasoTimeout:
            elapsed = time.perf_counter() - t0
            print(f"    TIMEOUT (>{TIMEOUT_CASO}s)  t={elapsed:.0f}s", flush=True)
            caso_resultado = {**caso, "phi_ib": None, "t_ib": round(elapsed, 3),
                              "particion_ib": f"TIMEOUT >{TIMEOUT_CASO}s"}
            prueba = caso["fila"] - 5
            escribir_resultado_excel(caso["fila"], f"TIMEOUT >{TIMEOUT_CASO}s", None, elapsed,
                                     prueba, caso["alc_str"], caso["mec_str"])

        except Exception as exc:
            print(f"    ERROR: {exc}", flush=True)
            caso_resultado = {**caso, "phi_ib": None, "t_ib": None,
                              "particion_ib": f"ERROR: {exc}"}
            prueba = caso["fila"] - 5
            escribir_resultado_excel(caso["fila"], f"ERROR: {exc}", None, None,
                                     prueba, caso["alc_str"], caso["mec_str"])

        # Upsert checkpoint
        idx = next((j for j, c in enumerate(resultados) if c["fila"] == caso["fila"]), None)
        if idx is not None:
            resultados[idx] = caso_resultado
        else:
            resultados.append(caso_resultado)
        guardar_checkpoint({"casos": resultados})

    elapsed_total = time.perf_counter() - t_total
    completados = [c for c in resultados if c.get("phi_ib") is not None]
    print(f"\n{'='*60}")
    print(f"  26A completado: {len(completados)}/{len(casos)} casos")
    print(f"  Tiempo total: {elapsed_total/3600:.2f}h  ({elapsed_total:.0f}s)")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
