"""Ejecuta Geometric para las 10 filas de 20A-Elementos que ya tienen QNodos.
Guarda fila a fila al Excel."""
import sys, time, gc
import numpy as np
import pandas as pd
import openpyxl
sys.path.insert(0, "/home/cami/Desktop/AnalisisDiseñoAlgoritmos/Proyecto/ProyectoAnalisis2026")

from src.strategies.geometric import Geometric

EXCEL = "DatosPruebas2026_1.xlsx"
SHEET = "20A-Elementos"
CSV   = "src/.samples/N20A.csv"
SISTEMA   = "ABCDEFGHIJKLMNOPQRST"
ESTADO    = "10000000000000000000"
CONDICION = "1" * 20
LOG       = "/tmp/geo_10rows.log"

# Geometric k=2..5: (col_partition, col_perdida, col_tiempo) 1-indexed openpyxl
GEO_COLS = {2: (7, 8, 9), 3: (13, 14, 15), 4: (19, 20, 21), 5: (25, 26, 27)}

FILAS = [
    (38, "ABDEGHJKMNPQST", "ABDEGHJKMNPQST"),
    (39, "ABDEGHJKMNPQST", "ACEGIKMOQS"),
    (40, "ABDEGHJKMNPQST", "BDFHJLNPRT"),
    (45, "ACEGIKMOQS",     "ABDEGHJKMNPQST"),
    (46, "ACEGIKMOQS",     "ACEGIKMOQS"),
    (47, "ACEGIKMOQS",     "BDFHJLNPRT"),
    (52, "BDFHJLNPRT",     "ABDEGHJKMNPQST"),
    (53, "BDFHJLNPRT",     "ACEGIKMOQS"),
    (54, "BDFHJLNPRT",     "BDFHJLNPRT"),
    (55, "BCDEFGJKLMNO",   "BCDEFGHIJKLMNO"),
]

def to_mask(sub):
    return "".join("1" if c in sub else "0" for c in SISTEMA)

def log(msg):
    print(msg, flush=True)

def guardar_fila(fila_idx, kdict):
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[SHEET]
    for k, (particion, perdida, tiempo) in kdict.items():
        cp, cl, ct = GEO_COLS[k]
        ws.cell(row=fila_idx, column=cp, value=particion)
        ws.cell(row=fila_idx, column=cl, value=perdida)
        ws.cell(row=fila_idx, column=ct, value=tiempo)
    wb.save(EXCEL)
    wb.close()

log("Cargando TPM (float32)...")
tpm = pd.read_csv(CSV, header=None).values.astype(np.float32)
log(f"TPM: {tpm.shape}")
geo = Geometric(tpm)
log(f"Geometric listo. Procesando {len(FILAS)} filas...\n")

t_total = time.perf_counter()

for fila_num, (fila_idx, alc_letras, mec_letras) in enumerate(FILAS, 1):
    alc_mask = to_mask(alc_letras)
    mec_mask = to_mask(mec_letras)
    n_max = max(len(alc_letras), len(mec_letras))
    log(f"[{fila_num:02d}/{len(FILAS)}] fila={fila_idx} n_max={n_max} | {alc_letras} / {mec_letras}")
    kdict = {}
    t_fila = time.perf_counter()
    for k in (2, 3, 4, 5):
        t0 = time.perf_counter()
        res = geo.aplicar_estrategia(
            estado_inicial=ESTADO,
            condicion=CONDICION,
            alcance=alc_mask,
            mecanismo=mec_mask,
            k=k,
        )
        elapsed_k = time.perf_counter() - t0
        perdida = float(res.perdida)
        particion = str(res.particion) if hasattr(res, "particion") else ""
        kdict[k] = (particion, perdida, round(elapsed_k, 4))
        log(f"  k={k} Geometric perdida={round(perdida,6)}  t={round(elapsed_k,4)}s")
        gc.collect()
    elapsed_fila = time.perf_counter() - t_fila
    guardar_fila(fila_idx, kdict)
    acum = (time.perf_counter() - t_total) / 60
    log(f"  -> [GUARDADO] fila {fila_idx} OK  {round(elapsed_fila,1)}s | acumulado {round(acum,1)}min\n")

log(f"Todas las filas completadas. Total: {round((time.perf_counter()-t_total)/60,1)} min")
