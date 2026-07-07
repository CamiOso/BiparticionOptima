"""Crea la hoja 28A-Elementos en DatosIBQNodos2026.xlsx.

Replica el patrón de 26A-Elementos: 7 grupos de alcance × 7 mecanismos + 1 especial = 50 casos.

Sistema n=28: ABCDEFGHIJKLMNOPQRSTUVWXYZ01
Estado inicial: 1000000000000000000000000000

Patrones de subconjuntos:
  mec/alc 28  = ABCDEFGHIJKLMNOPQRSTUVWXYZ01   (completo)
  mec/alc 27a = ABCDEFGHIJKLMNOPQRSTUVWXYZ0    (sin '1')
  mec/alc 27b = BCDEFGHIJKLMNOPQRSTUVWXYZ01    (sin 'A')
  mec/alc 26  = BCDEFGHIJKLMNOPQRSTUVWXYZ0     (sin 'A' ni '1')
  mec/alc 19  = ABDEGHJKMNPQSTVWYZ1            (tomar 2, saltar 1)
  mec/alc 14a = ACEGIKMOQSUWY0                 (posiciones impares)
  mec/alc 14b = BDFHJLNPRTVXZ1                (posiciones pares)

Uso:
    python scripts/crear_hoja_28A.py
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

EXCEL = Path("DatosIBQNodos2026.xlsx")
NOMBRE_HOJA = "28A-Elementos"

SISTEMA  = "ABCDEFGHIJKLMNOPQRSTUVWXYZ01"
ESTADO   = "1" + "0" * 27

# ── Subconjuntos canónicos ──────────────────────────────────────────────
S28  = "ABCDEFGHIJKLMNOPQRSTUVWXYZ01"
S27a = "ABCDEFGHIJKLMNOPQRSTUVWXYZ0"    # sin '1'
S27b = "BCDEFGHIJKLMNOPQRSTUVWXYZ01"    # sin 'A'
S26  = "BCDEFGHIJKLMNOPQRSTUVWXYZ0"     # sin 'A' ni '1'
# Patrón tomar-2-saltar-1: AB_DE_GH_JK_MN_PQ_ST_VW_YZ_1
S19  = "ABDEGHJKMNPQSTVWYZ1"            # 19 nodos
# Posiciones impares (A=1,C=3,...,0=27): 14 nodos
S14a = "ACEGIKMOQSUWY0"
# Posiciones pares  (B=2,D=4,...,1=28): 14 nodos
S14b = "BDFHJLNPRTVXZ1"
# Especial: 24 nodos (sin B, U, W, Y del sistema completo)
S24  = "ACDEFGHIJKLMNOPQRSTVXZ01"

# ── Grupos: (alcance, [mec1, mec2, ...]) ───────────────────────────────
ALCANCES = [
    (S28,  [S28, S27a, S27b, S26, S19, S14a, S14b]),
    (S27a, [S28, S27a, S27b, S26, S19, S14a, S14b]),
    (S27b, [S28, S27a, S27b, S26, S19, S14a, S14b]),
    (S26,  [S28, S27a, S27b, S26, S19, S14a, S14b]),
    (S19,  [S28, S27a, S27b, S26, S19, S14a, S14b]),
    (S14a, [S28, S27a, S27b, S26, S19, S14a, S14b]),
    (S14b, [S28, S27a, S27b, S26, S19, S14a, S14b]),
]
ESPECIAL = [(S24, S24)]

# ── Estilos ────────────────────────────────────────────────────────────
AZUL    = PatternFill("solid", fgColor="2E75B6")
CELESTE = PatternFill("solid", fgColor="BDD7EE")
GRIS    = PatternFill("solid", fgColor="F2F2F2")
BLANCO  = PatternFill("solid", fgColor="FFFFFF")

def bold_white(size=11):
    return Font(bold=True, color="FFFFFF", size=size)

def bold_blue(size=10):
    return Font(bold=True, color="1F3864", size=size)

center = Alignment(horizontal="center", vertical="center")
left   = Alignment(horizontal="left",   vertical="center")


def escribir_hoja(wb):
    if NOMBRE_HOJA in wb.sheetnames:
        del wb[NOMBRE_HOJA]
    ws = wb.create_sheet(NOMBRE_HOJA)

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12

    def meta(row, label, value):
        a = ws.cell(row, 1, label)
        a.font = bold_blue(); a.fill = CELESTE; a.alignment = left
        b = ws.cell(row, 2, value)
        b.font = Font(size=10); b.fill = CELESTE; b.alignment = left
        ws.merge_cells(f"B{row}:F{row}")

    meta(1, "Estado inicial", ESTADO)
    meta(2, "Sistema:", SISTEMA)
    ws.cell(3, 1, "Sistema Candidato:").font = bold_blue()
    ws.cell(3, 1).fill = CELESTE; ws.cell(3, 1).alignment = left
    ws.cell(3, 2, SISTEMA).font = Font(size=10)
    ws.cell(3, 2).fill = CELESTE; ws.cell(3, 2).alignment = left
    ws.merge_cells("B3:C3")

    c = ws.cell(3, 4, "PRUEBAS  BIPARTICIONES")
    c.font = bold_white(); c.fill = AZUL; c.alignment = center
    ws.merge_cells("D3:F3")

    c = ws.cell(4, 4, "IBQNodos")
    c.font = bold_white(); c.fill = AZUL; c.alignment = center
    ws.merge_cells("D4:F4")

    headers = ["#Prueba", "Alcance o Purview (t+1)", "Mecanismo(t)",
               "Partición", "Pérdida", "Tiempo"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(5, col, h)
        c.font = bold_white(); c.fill = AZUL; c.alignment = center
    ws.row_dimensions[5].height = 20

    fila = 6
    prueba = 1

    for alc_str, mecanismos in ALCANCES:
        for i, mec_str in enumerate(mecanismos):
            bg = GRIS if i % 2 == 0 else BLANCO
            ws.cell(fila, 1, prueba).fill = bg
            ws.cell(fila, 2, alc_str).fill = bg
            ws.cell(fila, 3, mec_str).fill = bg
            for col in [4, 5, 6]:
                ws.cell(fila, col).fill = bg
            for col in range(1, 7):
                ws.cell(fila, col).alignment = left
                ws.cell(fila, col).font = Font(size=10)
            fila += 1
            prueba += 1

    alc_str, mec_str = ESPECIAL[0]
    bg = CELESTE
    ws.cell(fila, 1, prueba).fill = bg; ws.cell(fila, 1).font = Font(size=10, bold=True)
    ws.cell(fila, 2, alc_str).fill = bg
    ws.cell(fila, 3, mec_str).fill = bg
    for col in [4, 5, 6]:
        ws.cell(fila, col).fill = bg

    print(f"Hoja '{NOMBRE_HOJA}' creada con {prueba} casos (filas 6–{fila}).")
    return ws


def main():
    wb = openpyxl.load_workbook(EXCEL)
    escribir_hoja(wb)
    wb.save(EXCEL)
    print(f"Guardado: {EXCEL.resolve()}")

    total = sum(len(m) for _, m in ALCANCES) + len(ESPECIAL)
    print(f"\nResumen 28A-Elementos:")
    print(f"  Total casos: {total}")
    for alc_str, mecanismos in ALCANCES:
        print(f"  alc={len(alc_str):2d} ({alc_str[:10]}...): {len(mecanismos)} mechs")
    print(f"  Especial alc={len(ESPECIAL[0][0])}: 1 caso")
    print(f"\nPatrones de subconjuntos usados:")
    for nombre, s in [("S28",S28),("S27a",S27a),("S27b",S27b),("S26",S26),
                       ("S19",S19),("S14a",S14a),("S14b",S14b),("S24",S24)]:
        print(f"  {nombre} ({len(s):2d} nodos): {s}")


if __name__ == "__main__":
    main()
