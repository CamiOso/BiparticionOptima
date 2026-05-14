"""Geometric para UNA fila. Lanzar 4 instancias en paralelo con &"""
import sys, time, gc, argparse
import numpy as np
import openpyxl

PROJECT = "/home/cami/Desktop/AnalisisDiseñoAlgoritmos/Proyecto/ProyectoAnalisis2026"
sys.path.insert(0, PROJECT)

EXCEL     = f"{PROJECT}/DatosPruebas2026_1.xlsx"
SHEET     = "20A-Elementos"
CSV       = f"{PROJECT}/src/.samples/N20A.npy"
SISTEMA   = "ABCDEFGHIJKLMNOPQRST"
ESTADO    = "10000000000000000000"
CONDICION = "1" * 20
GEO_COLS  = {2: (7,8,9), 3: (13,14,15), 4: (19,20,21), 5: (25,26,27)}

FILAS = {
    45: ("ACEGIKMOQS",   "ABDEGHJKMNPQST"),
    47: ("ACEGIKMOQS",   "BDFHJLNPRT"),
    52: ("BDFHJLNPRT",   "ABDEGHJKMNPQST"),
    53: ("BDFHJLNPRT",   "ACEGIKMOQS"),
    54: ("BDFHJLNPRT",   "BDFHJLNPRT"),
    55: ("BCDEFGJKLMNO", "BCDEFGHIJKLMNO"),
}

def to_mask(sub):
    return "".join("1" if c in sub else "0" for c in SISTEMA)

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("fila", type=int)
    args = parser.parse_args()
    fila_idx = args.fila
    alc_letras, mec_letras = FILAS[fila_idx]

    from src.strategies.geometric import Geometric
    print(f"[fila={fila_idx}] Cargando TPM...", flush=True)
    tpm = np.load(CSV)
    geo = Geometric(tpm)
    geo._usar_paralelizacion_costos = False

    alc_mask = to_mask(alc_letras)
    mec_mask = to_mask(mec_letras)
    n_max = max(len(alc_letras), len(mec_letras))
    print(f"[fila={fila_idx}] n_max={n_max} | {alc_letras} / {mec_letras}", flush=True)

    kdict = {}
    t_fila = time.perf_counter()
    for k in (2, 3, 4, 5):
        t0 = time.perf_counter()
        res = geo.aplicar_estrategia(
            estado_inicial=ESTADO, condicion=CONDICION,
            alcance=alc_mask, mecanismo=mec_mask, k=k,
        )
        elapsed = time.perf_counter() - t0
        kdict[k] = (str(res.particion) if hasattr(res, "particion") else "",
                    float(res.perdida), round(elapsed, 4))
        print(f"  [fila={fila_idx} k={k}] perdida={round(float(res.perdida),6)} t={round(elapsed,2)}s", flush=True)
        gc.collect()

    total = round(time.perf_counter() - t_fila, 1)
    # Guardar con un pequeño delay aleatorio para evitar colisiones
    import random, time as _time
    _time.sleep(random.uniform(0, 2))
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[SHEET]
    for k, (part, perd, t) in kdict.items():
        cp, cl, ct = GEO_COLS[k]
        ws.cell(row=fila_idx, column=cp, value=part)
        ws.cell(row=fila_idx, column=cl, value=perd)
        ws.cell(row=fila_idx, column=ct, value=t)
    wb.save(EXCEL)
    wb.close()
    print(f"  [fila={fila_idx}] GUARDADO {total}s", flush=True)
