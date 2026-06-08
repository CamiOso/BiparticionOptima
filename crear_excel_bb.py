"""Genera DatosBranchBound2026.xlsx a partir de DatosPruebas2026_1.xlsx.

Estructura de columnas por hoja (k=2,3,4,5):

  Para cada bloque k (k=2..5) se repite:
    QNodos   — Partición, Pérdida, Tiempo  (copiado del original)
    Geometric— Partición, Pérdida, Tiempo  (copiado del original)
    BranchBound — Partición, Pérdida, Tiempo  (VACÍO, para rellenar)

  Layout completo (27 columnas de datos + 3 de identificación = 30):
    1   #Prueba
    2   Alcance
    3   Mecanismo
    --- k=2 (cols 4-12) ---
    4-6   QNodos k=2
    7-9   Geometric k=2
    10-12 BranchBound k=2
    --- k=3 (cols 13-21) ---
    13-15 QNodos k=3
    16-18 Geometric k=3
    19-21 BranchBound k=3
    --- k=4 (cols 22-30) ---
    22-24 QNodos k=4
    25-27 Geometric k=4
    28-30 BranchBound k=4
    --- k=5 (cols 31-39) ---
    31-33 QNodos k=5
    34-36 Geometric k=5
    37-39 BranchBound k=5
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

ORIGEN = "DatosPruebas2026_1.xlsx"
DESTINO = "DatosBranchBound2026.xlsx"

HOJAS = [
    ("10A-Elementos",   "1000000000",                    "ABCDEFGHIJ",               "10A"),
    ("15B-Elementos",   "100000000000000",                "ABCDEFGHIJKLMNO",          "15B"),
    ("20A-Elementos",   "10000000000000000000",            "ABCDEFGHIJKLMNOPQRST",     "20A"),
    ("22A-Elementos",   "1000000000000000000000",          "ABCDEFGHIJKLMNOPQRSTUV",   "22A"),
    ("25A-Elementos ",  "1000000000000000000000000",       "ABCDEFGHIJKLMNOPQRSTUVWXY","25A"),
]

# Columnas de inicio en el ORIGINAL para cada k (1-indexed)
# QNodos k=2→cols 4,5,6 | Geometric k=2→cols 7,8,9 | etc.
ORIG_COLS = {
    2: {"q": (4, 5, 6), "g": (7, 8, 9)},
    3: {"q": (10, 11, 12), "g": (13, 14, 15)},
    4: {"q": (16, 17, 18), "g": (19, 20, 21)},
    5: {"q": (22, 23, 24), "g": (25, 26, 27)},
}

# Columnas de inicio en el DESTINO por k (9 cols por bloque: Q×3 + G×3 + BB×3)
# k=2 → cols 4-12, k=3 → cols 13-21, k=4 → cols 22-30, k=5 → cols 31-39
DST_START = {2: 4, 3: 13, 4: 22, 5: 31}

COLOR_QNODOS   = "D9EAD3"
COLOR_GEO      = "CFE2F3"
COLOR_BB       = "FFF2CC"
COLOR_HEADER_K = "434343"


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _bold(size: int = 11, color: str = "000000") -> Font:
    return Font(bold=True, size=size, color=color)


def crear_hoja(tag: str, nombre_src: str, estado: str, sistema: str):
    ws_src = wb_src[nombre_src]
    ws = wb_dst.create_sheet(title=tag)

    # ── Filas de encabezado ──────────────────────────────────────────────────
    ws.cell(1, 1, "Estado inicial").font = _bold()
    ws.cell(1, 2, estado)
    ws.cell(2, 1, "Sistema:").font = _bold()
    ws.cell(2, 2, sistema)
    ws.cell(3, 1, "Sistema Candidato:").font = _bold()
    ws.cell(3, 2, sistema)
    ws.cell(4, 2, "Subsistema").font = _bold()

    for k in [2, 3, 4, 5]:
        cs = DST_START[k]
        # fila 3: bloque k
        lbl = f"PRUEBAS  {k}-PARTICIONES"
        ws.cell(3, cs, lbl).font = _bold(color="FFFFFF")
        ws.cell(3, cs).fill = _fill(COLOR_HEADER_K)
        ws.merge_cells(start_row=3, start_column=cs, end_row=3, end_column=cs + 8)
        ws.cell(3, cs).alignment = Alignment(horizontal="center")

        # fila 4: nombre estrategia (3 bloques de 3 cols cada uno)
        for offset, label, color in [(0, "QNodos", COLOR_QNODOS),
                                      (3, "Geometric", COLOR_GEO),
                                      (6, "BranchBound", COLOR_BB)]:
            c = cs + offset
            ws.cell(4, c, label).font = _bold()
            ws.cell(4, c).fill = _fill(color)
            ws.merge_cells(start_row=4, start_column=c, end_row=4, end_column=c + 2)
            ws.cell(4, c).alignment = Alignment(horizontal="center")

        # fila 5: cabecera de columnas
        for offset, color in [(0, COLOR_QNODOS), (3, COLOR_GEO), (6, COLOR_BB)]:
            for i, lbl_col in enumerate(["Partición", "Pérdida", "Tiempo"]):
                cell = ws.cell(5, cs + offset + i, lbl_col)
                cell.font = _bold()
                cell.fill = _fill(color)
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Fila 5 para las columnas de identificación
    for c, lbl in [(1, "#Prueba"), (2, "Alcance o Purview (t+1)"), (3, "Mecanismo(t)")]:
        ws.cell(5, c, lbl).font = _bold()

    # ── Filas de datos ──────────────────────────────────────────────────────
    # Lee todas las filas de origen una vez
    datos_src = list(ws_src.iter_rows(min_row=6, values_only=True))

    fila_dst = 6
    for fila_src in datos_src:
        alc = fila_src[1]
        mec = fila_src[2]
        if alc is None:
            continue

        ws.cell(fila_dst, 2, alc)
        ws.cell(fila_dst, 3, mec)

        for k in [2, 3, 4, 5]:
            cs = DST_START[k]
            q_cols = ORIG_COLS[k]["q"]  # 1-indexed en original
            g_cols = ORIG_COLS[k]["g"]

            # QNodos (cols 4..6 del bloque destino)
            q_part = fila_src[q_cols[0] - 1]
            q_perd = fila_src[q_cols[1] - 1]
            q_t    = fila_src[q_cols[2] - 1]
            if q_perd is not None:
                ws.cell(fila_dst, cs,     q_part)
                ws.cell(fila_dst, cs + 1, round(float(q_perd), 6))
                ws.cell(fila_dst, cs + 2, q_t)

            # Geometric (cols 7..9 del bloque destino)
            g_part = fila_src[g_cols[0] - 1]
            g_perd = fila_src[g_cols[1] - 1]
            g_t    = fila_src[g_cols[2] - 1]
            if g_perd is not None:
                ws.cell(fila_dst, cs + 3, g_part)
                ws.cell(fila_dst, cs + 4, round(float(g_perd), 6))
                ws.cell(fila_dst, cs + 5, g_t)
            # BranchBound (cols 10..12 del bloque destino) — VACÍO

        fila_dst += 1

    # ── Ancho de columnas ────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    from openpyxl.utils import get_column_letter
    for k in [2, 3, 4, 5]:
        cs = DST_START[k]
        for offset in [0, 3, 6]:       # col partición de cada estrategia
            ws.column_dimensions[get_column_letter(cs + offset)].width = 65
        for offset in [1, 2, 4, 5, 7, 8]:  # pérdida y tiempo
            ws.column_dimensions[get_column_letter(cs + offset)].width = 13

    print(f"  {tag}: {fila_dst - 6} filas de datos copiadas")


wb_src = openpyxl.load_workbook(ORIGEN, data_only=True)
wb_dst = openpyxl.Workbook()
wb_dst.remove(wb_dst.active)

print(f"Creando {DESTINO}...")
for nombre_src, estado, sistema, tag in HOJAS:
    crear_hoja(tag, nombre_src, estado, sistema)

wb_dst.save(DESTINO)
print(f"Guardado: {DESTINO}")
