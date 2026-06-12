"""Genera DatosML2026.xlsx con resultados de las estrategias ML/alternativas.

Estrategias evaluadas:
    Louvain, InformacionBottleneck, ParticionHiperbolica,
    ParticionVariacional, REMCMC

Sistemas: 4A (4 nodos), 10A (10 nodos), 15B (15 nodos)
K values: 2, 3, 4, 5

Ejecutar:
    python crear_excel_ml.py
"""

from __future__ import annotations

import time
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# Configuración
# ──────────────────────────────────────────────────────────────────────────────

DESTINO = "DatosML2026.xlsx"
K_VALUES = [2, 3, 4, 5]

ESTRATEGIAS = [
    ("Louvain",        "src.estrategias.louvain",              "Louvain"),
    ("Info Bottleneck","src.estrategias.informacion_bottleneck","InformacionBottleneck"),
    ("Hiperbolica",    "src.estrategias.hiperbolica",           "ParticionHiperbolica"),
    ("Variacional",    "src.estrategias.variacional",           "ParticionVariacional"),
    ("REMCMC",         "src.estrategias.remcmc",                "REMCMC"),
]

# Sistema → (estado, sistema_letras, csv_path, casos[(alcance, mecanismo)])
SISTEMAS = {
    "4A": {
        "estado":  "1000",
        "sistema": "ABCD",
        "tpm":     None,   # se carga desde Gestor
        "casos": [
            ("ABCD", "ABCD"),
            ("ABCD", "ABC"),
            ("ABCD", "BCD"),
            ("ABC",  "ABCD"),
        ],
    },
    "10A": {
        "estado":  "1000000000",
        "sistema": "ABCDEFGHIJ",
        "tpm":     "src/.samples/N10A.csv",
        "casos": [
            ("ABCDEFGHIJ", "ABCDEFGHIJ"),
            ("ABCDEFGHIJ", "ABCDEFGHI"),
            ("ABCDEFGHIJ", "BCDEFGHIJ"),
            ("ABCDEFGHI",  "ABCDEFGHIJ"),
        ],
    },
    "15B": {
        "estado":  "100000000000000",
        "sistema": "ABCDEFGHIJKLMNO",
        "tpm":     "src/.samples/N15B.csv",
        "casos": [
            ("ABCDEFGHIJKLMNO", "ABCDEFGHIJKLMNO"),
            ("ABCDEFGHIJKLMNO", "ABCDEFGHIJKLMN"),
            ("ABCDEFGHIJKLMNO", "BCDEFGHIJKLMNO"),
        ],
    },
}

# ──────────────────────────────────────────────────────────────────────────────
# Colores
# ──────────────────────────────────────────────────────────────────────────────
COLORES_EST = {
    "Louvain":         "D9EAD3",
    "Info Bottleneck": "CFE2F3",
    "Hiperbolica":     "FFF2CC",
    "Variacional":     "F4CCCC",
    "REMCMC":          "E6D0DE",
}
COLOR_K_HEADER = {2: "434343", 3: "666666", 4: "999999", 5: "BBBBBB"}
COLOR_HEADER   = "1F4E79"


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)

def _font(bold=False, size=11, color="000000", italic=False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _thin_border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


# ──────────────────────────────────────────────────────────────────────────────
# Cargar TPMs
# ──────────────────────────────────────────────────────────────────────────────

def _cargar_tpm(nombre: str) -> np.ndarray:
    cfg = SISTEMAS[nombre]
    if cfg["tpm"] is None:
        from src.modelos.base.aplicacion import aplicacion
        from src.controladores.gestor import Gestor
        aplicacion.set_pagina_red_muestra("A")
        return Gestor(cfg["estado"]).cargar_red()
    return np.genfromtxt(cfg["tpm"], delimiter=",").astype(np.float32)


def _importar_estrategia(modulo: str, clase: str):
    import importlib
    mod = importlib.import_module(modulo)
    return getattr(mod, clase)


def _mascara(letras: str, sistema: str) -> str:
    return "".join("1" if c in letras else "0" for c in sistema)


# ──────────────────────────────────────────────────────────────────────────────
# Hoja plataformas
# ──────────────────────────────────────────────────────────────────────────────

def _crear_hoja_plataformas(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet(title="plataformas", index=0)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80

    # Fila 1 — título
    ws.merge_cells("A1:B1")
    c = ws.cell(1, 1, "Características plataforma")
    c.font = _font(bold=True, size=13, color="FFFFFF")
    c.fill = _fill(COLOR_HEADER)
    c.alignment = _center()
    ws.row_dimensions[1].height = 22

    # Fila 2 — encabezados columnas
    for col, texto in [(1, "Campo"), (2, "Valor")]:
        c = ws.cell(2, col, texto)
        c.font = _font(bold=True, color="FFFFFF")
        c.fill = _fill("434343")
        c.alignment = _center()

    # Datos de la plataforma
    filas = [
        ("Procesador",
         "Intel Core i7-6500U @ 2.50 GHz (Turbo 3.10 GHz)"),
        ("Características del procesador",
         "2 núcleos físicos, 4 hilos lógicos (Hyper-Threading); "
         "caché L1d 64 KiB / L1i 64 KiB / L2 512 KiB / L3 4 MiB; "
         "SSE4.2, AVX2; velocidad nominal 2.50 GHz, turbo hasta 3.10 GHz"),
        ("RAM",
         "~16 GB"),
        ("Sistema Operativo",
         "Debian GNU/Linux 12 (bookworm) — Kernel 6.6.119-antix.1-amd64-smp"),
        ("Python",
         "3.11.2"),
        ("NumPy",
         "2.4.4 (OpenBLAS + LAPACK)"),
        ("openpyxl",
         "3.1.5"),
        ("Software especializado",
         "Python 3.11.2 | NumPy 2.4.4 (OpenBLAS) | openpyxl 3.1.5 | "
         "python-pptx 1.0.2 | matplotlib | networkx"),
    ]

    for i, (campo, valor) in enumerate(filas, start=3):
        bg = "F4F6F8" if i % 2 == 0 else "FFFFFF"
        c_campo = ws.cell(i, 1, campo)
        c_campo.font = _font(bold=True)
        c_campo.fill = _fill(bg)
        c_campo.alignment = Alignment(vertical="center", wrap_text=True)
        c_campo.border = _thin_border()

        c_valor = ws.cell(i, 2, valor)
        c_valor.fill = _fill(bg)
        c_valor.alignment = Alignment(vertical="center", wrap_text=True)
        c_valor.border = _thin_border()
        ws.row_dimensions[i].height = 30


# ──────────────────────────────────────────────────────────────────────────────
# Construcción de la hoja
# ──────────────────────────────────────────────────────────────────────────────

def _crear_hoja(wb: openpyxl.Workbook, nombre_sistema: str, tpm: np.ndarray) -> None:
    cfg    = SISTEMAS[nombre_sistema]
    estado = cfg["estado"]
    sistema = cfg["sistema"]
    casos  = cfg["casos"]

    ws = wb.create_sheet(title=f"{nombre_sistema}-Elementos")

    # ── Fila 1-2: info del sistema ────────────────────────────────────────
    ws.cell(1, 1, "Estado inicial").font = _font(bold=True)
    ws.cell(1, 2, estado)
    ws.cell(2, 1, "Sistema:").font = _font(bold=True)
    ws.cell(2, 2, sistema)
    ws.cell(3, 1, "Estrategias:").font = _font(bold=True)
    ws.cell(3, 2, "Louvain | Info Bottleneck | Hiperbolica | Variacional | REMCMC")

    # ── Calcular columnas ─────────────────────────────────────────────────
    # Layout: col 1=#Prueba, col 2=Alcance, col 3=Mecanismo
    # Luego por cada k: 5 estrategias × 3 cols (Partición, Pérdida, Tiempo)
    # = 15 cols por bloque de k → 4 bloques = 60 cols + 3 = 63 total
    COL_START = 4   # primera columna de datos

    # Fila 5: encabezados de k (merged)
    # Fila 6: nombre de estrategia (merged 3 cols)
    # Fila 7: Partición | Pérdida | Tiempo

    # Encabezados fijos
    for r in range(5, 8):
        ws.cell(r, 1, "#Prueba" if r == 7 else None)
        ws.cell(r, 2, "Alcance" if r == 7 else ("Subsistema" if r == 6 else None))
        ws.cell(r, 3, "Mecanismo" if r == 7 else None)

    ws.cell(5, 1).font = _font(bold=True)
    ws.cell(7, 1).font = _font(bold=True)
    ws.cell(7, 2).font = _font(bold=True)
    ws.cell(7, 3).font = _font(bold=True)
    for c in range(1, 4):
        ws.cell(7, c).fill = _fill(COLOR_HEADER)
        ws.cell(7, c).font = _font(bold=True, color="FFFFFF")
        ws.cell(7, c).alignment = _center()

    col = COL_START
    for k in K_VALUES:
        k_start = col
        k_end   = col + len(ESTRATEGIAS) * 3 - 1

        # Fila 5: encabezado k (merged)
        ws.merge_cells(
            start_row=5, start_column=k_start,
            end_row=5, end_column=k_end
        )
        cell_k = ws.cell(5, k_start, f"k = {k}")
        cell_k.fill  = _fill(COLOR_K_HEADER[k])
        cell_k.font  = _font(bold=True, size=13, color="FFFFFF")
        cell_k.alignment = _center()

        for nombre_est, _, _ in ESTRATEGIAS:
            e_start = col
            # Fila 6: nombre estrategia (merged 3 cols)
            ws.merge_cells(
                start_row=6, start_column=col,
                end_row=6, end_column=col + 2
            )
            cell_e = ws.cell(6, col, nombre_est)
            cell_e.fill  = _fill(COLORES_EST[nombre_est])
            cell_e.font  = _font(bold=True)
            cell_e.alignment = _center()

            # Fila 7: subencabezados
            for i, sub in enumerate(["Partición", "Pérdida", "Tiempo (s)"]):
                c = ws.cell(7, col + i, sub)
                c.fill = _fill(COLORES_EST[nombre_est])
                c.font = _font(bold=True, size=10)
                c.alignment = _center()

            col += 3

    # Anchos de columna
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    for i in range(COL_START, col):
        letra = get_column_letter(i)
        # Partición es ancho, Pérdida y Tiempo son angostos
        pos = (i - COL_START) % 3
        ws.column_dimensions[letra].width = 35 if pos == 0 else 10

    # Congelar primeras 3 cols y 7 filas
    ws.freeze_panes = "D8"

    # ── Filas de datos ────────────────────────────────────────────────────
    clases_cargadas = {
        nombre: _importar_estrategia(mod, cls)
        for nombre, mod, cls in ESTRATEGIAS
    }

    condicion = "1" * len(sistema)

    for prueba_idx, (alc_letras, mec_letras) in enumerate(casos, start=1):
        fila = 7 + prueba_idx
        alc_mask = _mascara(alc_letras, sistema)
        mec_mask = _mascara(mec_letras, sistema)

        ws.cell(fila, 1, prueba_idx).alignment = _center()
        ws.cell(fila, 2, alc_letras)
        ws.cell(fila, 3, mec_letras)

        print(f"\n  [{nombre_sistema}] Caso {prueba_idx}/{len(casos)}: "
              f"alc={alc_letras} | mec={mec_letras}")

        col = COL_START
        for nombre_est, _, _ in ESTRATEGIAS:
            ClaseEst = clases_cargadas[nombre_est]
            for k in K_VALUES:
                try:
                    instancia = ClaseEst(tpm)
                    t0 = time.perf_counter()
                    res = instancia.aplicar_estrategia(
                        estado_inicial=estado,
                        condicion=condicion,
                        alcance=alc_mask,
                        mecanismo=mec_mask,
                        k=k,
                    )
                    elapsed = time.perf_counter() - t0
                    particion = str(res.particion)
                    perdida   = round(float(res.perdida), 6)
                    print(f"    {nombre_est} k={k}: φ={perdida:.4f}  t={elapsed:.2f}s")
                except Exception as e:
                    particion = f"ERROR: {e}"
                    perdida   = None
                    elapsed   = None
                    print(f"    {nombre_est} k={k}: ERROR — {e}")

                # Buscar la columna correcta para esta estrategia+k
                # col_est = COL_START + k_idx*15 + est_idx*3
                k_idx  = K_VALUES.index(k)
                e_idx  = [n for n, _, _ in ESTRATEGIAS].index(nombre_est)
                c_part = COL_START + k_idx * len(ESTRATEGIAS) * 3 + e_idx * 3
                c_perd = c_part + 1
                c_time = c_part + 2

                color = COLORES_EST[nombre_est]
                for c_col, valor in [(c_part, particion), (c_perd, perdida), (c_time, elapsed)]:
                    cell = ws.cell(fila, c_col, valor)
                    cell.fill = _fill(color)
                    cell.alignment = _center() if c_col != c_part else Alignment(
                        horizontal="left", vertical="center", wrap_text=True)
                    cell.border = _thin_border()

            col += 3  # no se usa directamente pero mantiene la lógica

    # Altura de filas de encabezado
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 30


# ──────────────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────────────

def main() -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # quitar hoja vacía por defecto

    t_total = time.perf_counter()

    _crear_hoja_plataformas(wb)
    print("  ✓ Hoja plataformas creada")

    for nombre_sistema in SISTEMAS:
        print(f"\n{'═'*60}")
        print(f"  Sistema: {nombre_sistema}")
        print(f"{'═'*60}")
        tpm = _cargar_tpm(nombre_sistema)
        print(f"  TPM cargada: {tpm.shape}")
        _crear_hoja(wb, nombre_sistema, tpm)
        wb.save(DESTINO)
        print(f"  ✓ Hoja {nombre_sistema} guardada → {DESTINO}")

    print(f"\n{'─'*60}")
    print(f"  Excel generado en: {DESTINO}")
    print(f"  Tiempo total: {time.perf_counter() - t_total:.1f} s")
    print(f"{'─'*60}\n")


if __name__ == "__main__":
    main()
