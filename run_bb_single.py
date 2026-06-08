"""BranchBound para UNA fila de DatosBranchBound2026.xlsx, k=2..5.

Uso:
    python run_bb_single.py <hoja> <fila>                        # k=2..5
    python run_bb_single.py <hoja> <fila> --start-k 3           # desde k=3
    python run_bb_single.py 10A 46 --umbral 14 --arranques 8

El argumento <hoja> es el nombre de la hoja en DatosBranchBound2026.xlsx.
El argumento <fila> es la fila Excel (≥6).
"""
import argparse
import gc
import os
import random
import sys
import time

import numpy as np
import openpyxl

PROJECT = "/home/cami/Desktop/AnalisisDiseñoAlgoritmos/Proyecto/ProyectoAnalisis2026"
sys.path.insert(0, PROJECT)

EXCEL = f"{PROJECT}/DatosBranchBound2026.xlsx"
LOCK  = EXCEL + ".lock"

HOJAS_CONFIG = {
    "10A": {"csv": f"{PROJECT}/src/.samples/N10A.csv",
            "sistema": "ABCDEFGHIJ",
            "estado":  "1000000000"},
    "15B": {"csv": f"{PROJECT}/src/.samples/N15B.csv",
            "sistema": "ABCDEFGHIJKLMNO",
            "estado":  "100000000000000"},
    "20A": {"npy": f"{PROJECT}/src/.samples/N20A.npy",
            "sistema": "ABCDEFGHIJKLMNOPQRST",
            "estado":  "10000000000000000000"},
    "22A": {"npy": f"{PROJECT}/src/.samples/N22A.npy",
            "sistema": "ABCDEFGHIJKLMNOPQRSTUV",
            "estado":  "1000000000000000000000"},
    "25A": {"npy": f"{PROJECT}/src/.samples/N25A.npy",
            "sistema": "ABCDEFGHIJKLMNOPQRSTUVWXY",
            "estado":  "1000000000000000000000000"},
}

# Columna de inicio de BranchBound para cada k en el Excel destino
# Layout: k=2→cols 4-12, k=3→cols 13-21, k=4→cols 22-30, k=5→cols 31-39
# Dentro de cada bloque: QNodos(3) + Geometric(3) + BB(3) = 9 cols
# BB ocupa los últimos 3 → offset 6 dentro del bloque
BB_COLS = {
    2: (4  + 6, 4  + 7, 4  + 8),   # cols 10, 11, 12
    3: (13 + 6, 13 + 7, 13 + 8),   # cols 19, 20, 21
    4: (22 + 6, 22 + 7, 22 + 8),   # cols 28, 29, 30
    5: (31 + 6, 31 + 7, 31 + 8),   # cols 37, 38, 39
}


def to_mask(sub: str, sistema: str) -> str:
    return "".join("1" if c in sub else "0" for c in sistema)


def cargar_tpm(cfg: dict) -> np.ndarray:
    if "npy" in cfg:
        return np.load(cfg["npy"], mmap_mode="r")
    return np.genfromtxt(cfg["csv"], delimiter=",")


def leer_fila(hoja: str, fila_idx: int) -> tuple[str, str]:
    wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)
    ws = wb[hoja]
    row = list(ws.iter_rows(min_row=fila_idx, max_row=fila_idx, values_only=True))[0]
    alc, mec = row[1], row[2]
    wb.close()
    return alc, mec


def guardar_k(hoja: str, fila_idx: int, k: int, part: str, perd: float, elapsed: float):
    col_p, col_e, col_t = BB_COLS[k]
    for _ in range(20):
        try:
            fd = os.open(LOCK, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.close(fd)
            break
        except FileExistsError:
            time.sleep(random.uniform(0.5, 2.0))
    else:
        raise RuntimeError("No se pudo adquirir lock del Excel")
    try:
        wb = openpyxl.load_workbook(EXCEL)
        ws = wb[hoja]
        ws.cell(row=fila_idx, column=col_p, value=part)
        ws.cell(row=fila_idx, column=col_e, value=round(perd, 6))
        ws.cell(row=fila_idx, column=col_t, value=round(elapsed, 4))
        wb.save(EXCEL)
        wb.close()
    finally:
        try:
            os.unlink(LOCK)
        except FileNotFoundError:
            pass


def ajustar_cache(mec_n: int):
    cn, cs = (512, 1024) if mec_n <= 13 else \
             (128, 512)  if mec_n <= 17 else \
             (64,  256)  if mec_n <= 20 else \
             (32,  128)
    import src.modelos.nucleo.ncubo   as _mn
    import src.modelos.nucleo.sistema as _ms
    _mn._MAX_MEMO_NCUBE    = cn
    _ms._MAX_MEMO_SISTEMA  = cs
    print(f"  cache ncube={cn} sistema={cs} (mec_n={mec_n})", flush=True)


if __name__ == "__main__":
    os.environ.setdefault("OMP_NUM_THREADS", "1")
    os.environ.setdefault("OPENBLAS_NUM_THREADS", "1")

    parser = argparse.ArgumentParser()
    parser.add_argument("hoja",  help="Hoja (10A, 15B, 20A, 22A, 25A)")
    parser.add_argument("fila",  type=int, help="Fila Excel ≥6")
    parser.add_argument("--start-k",   type=int, default=2, choices=[2, 3, 4, 5])
    parser.add_argument("--umbral",    type=int, default=14)
    parser.add_argument("--arranques", type=int, default=8)
    parser.add_argument("--hamming",   type=int, default=3)
    parser.add_argument("--pasos-sa",  type=int, default=800)
    parser.add_argument("--max-exact-k", type=int, default=200_000)
    args = parser.parse_args()

    hoja     = args.hoja
    fila_idx = args.fila

    if hoja not in HOJAS_CONFIG:
        print(f"ERROR: hoja '{hoja}' no reconocida. Opciones: {list(HOJAS_CONFIG)}", file=sys.stderr)
        sys.exit(1)

    cfg     = HOJAS_CONFIG[hoja]
    sistema = cfg["sistema"]
    estado  = cfg["estado"]
    cond    = "1" * len(sistema)

    print(f"[{hoja} fila={fila_idx}] Leyendo alcance/mecanismo...", flush=True)
    alc_letras, mec_letras = leer_fila(hoja, fila_idx)
    if not alc_letras or not mec_letras:
        print(f"[{hoja} fila={fila_idx}] Fila vacía, saltando.", flush=True)
        sys.exit(0)

    alc_mask = to_mask(alc_letras, sistema)
    mec_mask = to_mask(mec_letras, sistema)
    print(f"[{hoja} fila={fila_idx}] alc={alc_letras[:25]} ({len(alc_letras)}) mec={mec_letras[:25]} ({len(mec_letras)})",
          flush=True)

    ajustar_cache(len(mec_letras))

    from src.estrategias.branch_bound import BranchBound
    print(f"[{hoja} fila={fila_idx}] Cargando TPM...", flush=True)
    tpm = cargar_tpm(cfg)

    bb = BranchBound(
        tpm,
        umbral_exacto=args.umbral,
        n_sa_arranques=args.arranques,
        radio_hamming=args.hamming,
        pasos_sa=args.pasos_sa,
        max_exact_k=args.max_exact_k,
    )

    n_total = len(alc_letras) + len(mec_letras)

    for k in (kk for kk in [2, 3, 4, 5] if kk >= args.start_k):
        if k == 2:
            modo = "EXACTO" if n_total <= args.umbral else f"HEURISTICO(n={n_total})"
        else:
            from src.estrategias.branch_bound import BranchBound as _BB
            s = _BB._stirling(n_total, k)
            modo = f"EXACTO(S={s:,})" if s <= args.max_exact_k else f"HEURISTICO(S={s:,})"

        t0 = time.perf_counter()
        res = bb.aplicar_estrategia(
            estado_inicial=estado,
            condicion=cond,
            alcance=alc_mask,
            mecanismo=mec_mask,
            k=k,
        )
        elapsed = time.perf_counter() - t0
        perdida = float(res.perdida)
        part    = str(res.particion)

        print(f"  [k={k} {modo}] pérdida={round(perdida,6)}  t={round(elapsed,2)}s", flush=True)
        guardar_k(hoja, fila_idx, k, part, perdida, elapsed)
        print(f"  [k={k}] GUARDADO ✓", flush=True)
        gc.collect()

    print(f"[{hoja} fila={fila_idx}] COMPLETO", flush=True)
